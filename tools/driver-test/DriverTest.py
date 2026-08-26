import os
import re
import openpyxl
import wmi
import glob
from openpyxl.styles import PatternFill

# 公共配置
EXCEL_NAME = 'driver_test.xlsx'
TARGET_DIR = "SU_LP88_Drivers_WIN11_X64"

# Excel样式
HEADER_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")


# ================== 公共函数 ==================
def get_sorted_subdirs(root_dir):
    """获取按数字排序的子目录列表"""
    dirs = []
    for name in os.listdir(root_dir):
        path = os.path.join(root_dir, name)
        if os.path.isdir(path):
            match = re.match(r'^(\d+)', name)
            dirs.append((
                int(match.group(1)) if match else float('inf'),
                path,
                name
            ))
    return [x[1] for x in sorted(dirs, key=lambda x: x[0])]


# ================== 第一部分功能：提取驱动名称 ==================
def extract_bat_driver_name(file_path):
    """从bat文件提取驱动名称"""
    pattern = r"=+\s*([^=]+?)\s+Install\s"
    try:
        # 尝试UTF-8解码
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except UnicodeDecodeError:
        # 回退到GBK解码
        try:
            with open(file_path, 'r', encoding='gbk') as f:
                content = f.read()
        except Exception as e:
            return f"解码失败: {str(e)}"
    except Exception as e:
        return f"读取失败: {str(e)}"

    match = re.search(pattern, content)
    return match.group(1).strip() if match else "未找到驱动名称"


def find_bat_files(dir_path):
    """递归查找bat文件"""
    # 优先查找Install.bat
    bat_path = os.path.join(dir_path, "Install.bat")
    if os.path.exists(bat_path):
        return [bat_path]

    # 其次查找Install_Apps.bat
    bat_path = os.path.join(dir_path, "Install_Apps.bat")
    if os.path.exists(bat_path):
        return [bat_path]

    # 递归查找子目录
    bat_files = []
    for entry in os.listdir(dir_path):
        full_path = os.path.join(dir_path, entry)
        if os.path.isdir(full_path):
            bat_files.extend(find_bat_files(full_path))
    return bat_files


# ================== 第二部分功能：驱动安装验证 ==================
def get_installed_drivers():
    """获取已安装驱动信息"""
    c = wmi.WMI()
    drivers = {}
    try:
        for driver in c.Win32_PnPSignedDriver():
            if driver.DeviceName and driver.DriverVersion:
                key = (
                    driver.DeviceName.strip().lower(),
                    driver.DriverVersion.strip().lower()
                )
                drivers[key] = {
                    'name': driver.DeviceName,
                    'version': driver.DriverVersion,
                    'manufacturer': driver.Manufacturer
                }
    except Exception as e:
        print(f"驱动信息获取失败: {e}")
    return drivers


def process_inf_file(inf_path, installed_drivers):
    """处理单个inf文件进行验证"""
    try:
        # 处理不同编码格式
        try:
            with open(inf_path, 'r', encoding='utf-16') as f:
                content = f.read()
        except:
            with open(inf_path, 'r', encoding='utf-8') as f:
                content = f.read()

        # 提取驱动版本
        version_match = re.search(
            r'DriverVer\s*=\s*\d+/\d+/\d+\s*,\s*([\d\.]+)',
            content,
            re.IGNORECASE
        )
        if not version_match:
            return "未找到版本信息"
        driver_version = version_match.group(1).lower()

        # 提取驱动名称
        strings_section = re.search(
            r'\[Strings\][^\[]+',
            content,
            re.DOTALL | re.IGNORECASE
        )
        if not strings_section:
            return "未找到名称信息"

        driver_names = re.findall(
            r'=\s*"(.*?)"',
            strings_section.group(),
            re.IGNORECASE
        )
        if not driver_names:
            return "未找到名称信息"

        # 验证驱动状态
        normalized_names = [name.lower() for name in driver_names]
        for key in installed_drivers:
            stored_ver = installed_drivers[key]['version'].lower()
            stored_name = installed_drivers[key]['name'].lower()

            if stored_ver == driver_version:
                if any(name in stored_name for name in normalized_names):
                    return "驱动验证通过"
        return "驱动验证不通过"
    except Exception as e:
        return f"文件处理错误: {str(e)}"


# ================== Excel操作 ==================
def init_excel():
    """初始化Excel文件和表头"""
    if os.path.exists(EXCEL_NAME):
        wb = openpyxl.load_workbook(EXCEL_NAME)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['B1'] = "提取的驱动名称"
        ws['E1'] = "安装验证结果"
        ws['B1'].fill = HEADER_FILL
        ws['E1'].fill = HEADER_FILL
        wb.save(EXCEL_NAME)
    return wb


def write_excel_column(column, data, header):
    """写入指定列数据"""
    wb = init_excel()
    ws = wb.active

    # 更新表头
    ws[f'{column}1'] = header
    ws[f'{column}1'].fill = HEADER_FILL

    # 写入数据
    for idx, value in enumerate(data, start=2):
        ws[f'{column}{idx}'] = value

    # 自动调整列宽
    max_len = max(len(str(v)) for v in data) if data else 0
    ws.column_dimensions[column].width = max(max_len + 2, len(header) + 2)

    wb.save(EXCEL_NAME)


# ================== 主流程 ==================
def main():
    # 获取排序后的子目录列表
    subdirs = get_sorted_subdirs(TARGET_DIR)

    # 第一部分：提取驱动名称
    bat_results = []
    for subdir in subdirs:
        bat_files = find_bat_files(subdir)
        if not bat_files:
            bat_results.append("未找到bat文件")
            continue

        names = []
        for bat_file in bat_files:
            name = extract_bat_driver_name(bat_file)
            if name not in names and "失败" not in name:
                names.append(name)
        bat_results.append('；'.join(names) if names else "名称提取失败")

    write_excel_column('B', bat_results, "提取的驱动名称")

    # 第二部分：验证驱动安装
    installed_drivers = get_installed_drivers()
    verify_results = []

    for subdir in subdirs:
        inf_files = glob.glob(os.path.join(subdir, '**/*.inf'), recursive=True)
        if not inf_files:
            verify_results.append("未找到inf文件")
            continue

        # 处理所有inf文件直到找到有效结果
        final_result = "未找到有效验证信息"
        for inf_file in inf_files:
            result = process_inf_file(inf_file, installed_drivers)
            if "通过" in result:
                final_result = result
                break
            elif "不通过" in result:
                final_result = result
        verify_results.append(final_result)

    write_excel_column('E', verify_results, "安装验证结果")


if __name__ == "__main__":
    main()