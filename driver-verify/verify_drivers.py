# -*- coding: utf-8 -*-
r"""
verify_drivers.py —— 驱动清单校验器（在你自己的机器上运行，配合 collect_drivers.ps1）
====================================================================================
输入：
  1) 驱动清单 Excel（MegaBook2_DriverList_*.xlsx，sheet「The lastest release」）
  2) 目标机采集出的 driver_collect.json（由 collect_drivers.ps1 生成）

判断（对应你提出的 5 点）：
  [检查1] 表格中的驱动是否都在（缺失 -> FAIL）
  [检查2] 表格驱动版本是否与实机一致（实机 == 期望 -> 一致PASS；实机 > 期望 -> 偏高PASS并输出；实机 < 期望 -> FAIL）
  [检查3] 表格中的驱动是否都已签名（未签名 -> FAIL）
  [检查4] 实机上所有驱动是否已签名（表内未签名 -> FAIL；表外未签名 -> 仅输出WARN，不算FAIL）
  [检查5] 实机是否有异常设备（黄感/未挂载/无驱动等 -> 输出，供人工判断）

用法：
  # 离线模式（推荐）：先在目标机跑 collect_drivers.ps1 拿到 json，再在本机比对
  D:\py311\python.exe verify_drivers.py --excel "清单.xlsx" --json driver_collect.json

  # 本机模式：直接对“当前这台机器”采集并比对（会调用同目录 collect_drivers.ps1）
  D:\py311\python.exe verify_drivers.py --excel "清单.xlsx" --local

  # 远程模式：通过 PowerShell Remoting(WinRM) 采集目标机（需目标机已 Enable-PSRemoting）
  D:\py311\python.exe verify_drivers.py --excel "清单.xlsx" --remote 192.168.1.50 --user Administrator
"""
import argparse
import ctypes
import json
import os
import re
import subprocess
import sys
from datetime import datetime


def is_admin():
    """当前进程是否具有管理员权限"""
    try:
        return ctypes.windll.shell32.IsUserAnAdmin() != 0
    except Exception:
        return False

# ------------------------------------------------------------------ 路径(兼容 PyInstaller onefile)

def res_dir():
    """资源目录：打包后取 _MEIPASS 临时解压目录（放捆绑的 collect_drivers.ps1）"""
    return getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))


def app_dir():
    """可写输出目录：打包后取 exe 所在目录（放 json / 报告），未打包取脚本目录"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


# ------------------------------------------------------------------ 工具函数

def norm(s):
    """名称规范化：小写、去除 ® ™ (R) (TM)、标点转空格、压缩空格。
    例：'Intel(R) Serial IO' -> 'intel serial io'；'SenaryAudio' -> 'senaryaudio'"""
    if s is None:
        return ""
    s = str(s)
    s = s.replace('®', ' ').replace('™', ' ')
    s = re.sub(r'\((r|tm)\)', ' ', s, flags=re.I)
    s = re.sub(r'[^0-9a-zA-Z一-鿿]+', ' ', s)  # 非字母数字汉字 -> 空格
    return re.sub(r'\s+', ' ', s).strip().lower()


def tokens(s):
    t = norm(s).split()
    return [x for x in t if len(x) > 1 or x.isdigit()]  # 丢掉单字母噪声(如 r)


def parse_ver(v):
    """把 '10.1.56.28' / '1.0.00000.222' 解析成整数元组，便于比较。"""
    if v is None:
        return None
    nums = re.findall(r'\d+', str(v))
    if not nums:
        return None
    return tuple(int(x) for x in nums)


def inf_version(p):
    """从INF包记录取可比较版本：Version 为正常点分号版本则用之；
    否则(旧采集器对扩展类INF会把 Extension ID 的GUID误存进 Version)从 Signer 里抽“日期后的版本号”。"""
    v = str(p.get('Version', ''))
    if re.search(r'\d+\.\d+', v):
        return v
    m = re.search(r'\d+\.\d+[\d.]*', str(p.get('Signer', '')))
    return m.group(0) if m else v


def cmp_ver(actual, expected):
    """比较版本。返回 'equal' / 'higher' / 'lower' / 'unknown'。补齐位数后按元组比较。"""
    a, e = parse_ver(actual), parse_ver(expected)
    if a is None or e is None:
        return 'unknown'
    n = max(len(a), len(e))
    a = a + (0,) * (n - len(a))
    e = e + (0,) * (n - len(e))
    if a == e:
        return 'equal'
    return 'higher' if a > e else 'lower'


# ------------------------------------------------------------------ 读取数据

def load_excel(path):
    """读取 sheet「The lastest release」，返回 [{name, vendor, version, remark}]。
    表头在第3行：B=Device Name, C=Vendor, D=Driver Version, H=Remark。数据从第4行起。"""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['The lastest release'] if 'The lastest release' in wb.sheetnames else wb.worksheets[0]
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        # 列: A(0) 空, B(1)=名称, C(2)=厂商, D(3)=版本, ... H(7)=备注
        name = r[1] if len(r) > 1 else None
        if not name or not str(name).strip():
            continue
        rows.append({
            'name': str(name).strip(),
            'vendor': (str(r[2]).strip() if len(r) > 2 and r[2] is not None else ''),
            'version': (str(r[3]).strip() if len(r) > 3 and r[3] is not None else ''),
            'remark': (str(r[7]).strip() if len(r) > 7 and r[7] is not None else ''),
        })
    return rows


def load_collect(path):
    with open(path, encoding='utf-8-sig') as f:   # 采集 json 带 BOM，必须 utf-8-sig
        return json.load(f)


def run_collector(ps1, out_json, remote=None, user=None):
    """调用 collect_drivers.ps1 生成 json。remote 为空则采本机，否则走 WinRM。"""
    if remote:
        # 远程：Invoke-Command 把脚本送到目标机执行，结果写回本地 out_json
        cmd = (
            f"$c = Get-Credential -UserName '{user or 'Administrator'}' -Message '目标机凭据';"
            f"Invoke-Command -ComputerName {remote} -Credential $c -FilePath '{ps1}' "
            f"-ArgumentList '{out_json}';"
        )
        # 注：远程执行时 ps1 内 $Out 写的是目标机路径，实际生产中更稳的做法是
        # 直接在目标机跑 ps1 再把 json 拷回来（离线模式）。远程模式此处仅作示例骨架。
        print("[提示] 远程模式依赖目标机已开启 WinRM，若失败请改用离线模式。")
        subprocess.run(["powershell", "-NoProfile", "-Command", cmd], check=False)
    else:
        subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass",
             "-File", ps1, "-Out", out_json],
            check=True)
    return load_collect(out_json)


# ------------------------------------------------------------------ 匹配

# 别名表：给“按设备名匹配不到”的表格项手动指路。
# key = 表格 Device Name（原样）；value = 正则列表，会在 设备名/厂商/INF原始名/发布名 里搜。
# 目标机采到真实数据后，把匹配不上的项补进来即可。
ALIASES = {
    # LunarLake 平台 Chipset 不产生名为 "Chipset" 的设备，而是 SMBus/SPI 等系统设备(INF=lunarlakesystem.inf)
    'Intel® Chipset': [r'chipset', r'arrowlake.*system', r'pantherlake.*system',
                       r'lunarlake.*system', r'smbus', r'spi.*flash'],
    # LunarLake 核显改名为 Intel(R) Arc(TM) xxxV GPU；PantherLake 为 Intel(R) Graphics；INF 均为 iigd_*.inf
    'Intel(R) UHD Graphics': [r'intel.*graphics$', r'\barc\b.*gpu', r'iigd'],
    'Intel(R) AI Boost': [r'\bnpu\b'],                        # AI Boost 在新平台改名为 Intel(R) NPU
    'SenaryAudio': [r'senary\s*audio$'],                      # 清单无空格，实机为 Senary Audio（勿匹配到 ...Audio Effects）
    # 中文系统里设备名本地化为「英特尔(R) 无线 Bluetooth(R)」(无线=Wireless)，INF=ibtusb.inf
    'Intel(R) Wireless Bluetooth®': [r'wireless.*bluetooth', r'无线\s*bluetooth', r'ibtusb'],
    # BE200 实为 Wi-Fi 7 网卡(实机名 Intel(R) Wi-Fi 7 BE200 320MHz)，清单写成 6E；INF=netwtw6e/netwtw08.inf
    'Intel(R) Wi-Fi 6E BE200': [r'be200', r'wi.?fi.*be200', r'netwtw6e', r'netwtw08'],
    'MEP Camera Opt In Ext Inf Installation': [r'mepoptinext'],  # 正解是 oem34/mepoptinext.inf@1.0.0.1(opt-in扩展外壳)，非 mep_camera_component(3.0.x)
    'Windows Studio Effects Camera': [r'studio\s*effects'],
    'Intel® Smart Sound Technology': [r'smart\s*sound'],
    'STK Ambient Light, Proximity Sensor': [r'ambient\s*light', r'stkw?3000', r'proximity'],
    'Windows Bosch Accelerometer Driver': [r'accelerometer', r'bosch'],
    'Intel(R) PPM Provisioning Package': [r'ppm\s*provisioning', r'\bppm\b'],
    'Intel(R) Integrated Sensor Solution': [r'ish(heci|oed)\.inf'],  # ISS=Intel Sensor Hub，核心驱动 ishheci.inf/ishoed.inf @ 5.8.62.0（.inf锚定排除扩展模板；不是 issei.inf 那个2543.x）
}


# 清单里有几项其实不是“驱动”而是【微软商店(UWP)应用】，不会出现在驱动/INF枚举里，
# 需用 Get-AppxPackage 采集的 apps 列表来核对(采集器已一并采集)。
# key = 清单 Device Name（原样）；value = 正则列表，在 Appx 包 Name 里搜。
STORE_APPS = {
    'Realtek Audio Control': [r'realtek.*audio'],          # 商店应用 RealtekSemiconductorCorp.RealtekAudioControl
    'intel GraphicsSoftware': [r'intel.*graphic'],         # 商店应用 Intel Graphics Software / Command Center
}


def check_store_app(expected, collect):
    """核对“商店(UWP)应用”类清单项。返回 driver_results 风格的 dict（含 overall）。
    - 采集数据无 apps 字段(旧版采集器)：WARN，提示用新版采集器重采。
    - apps 里按正则找到：比对版本(equal/higher=PASS，lower=FAIL)。
    - 找不到：FAIL(未安装)。"""
    base = {'name': expected['name'], 'vendor': expected['vendor'],
            'expected': expected['version'], 'sign': 'N/A(应用)'}
    apps = collect.get('apps')
    if apps is None:
        base.update({'actual': '', 'method': 'STORE_APP(本次未采集应用清单)', 'ver': '待采集',
                     'overall': 'WARN',
                     'note': '此项为微软商店(UWP)应用，非驱动；本次采集数据未含应用清单，请用新版采集器重采后再判定'})
        return base
    if isinstance(apps, dict):
        apps = [apps]
    rgxs = [re.compile(p, re.I) for p in STORE_APPS[expected['name']]]
    hits = [a for a in apps if any(r.search(a.get('Name', '') or '') for r in rgxs)]
    if not hits:
        base.update({'actual': '', 'method': 'STORE_APP', 'ver': '未安装', 'overall': 'FAIL',
                     'note': '微软商店(UWP)应用未安装'})
        return base
    vers = sorted({str(a.get('Version', '')) for a in hits})
    cmps = [cmp_ver(a.get('Version'), expected['version']) for a in hits]
    rank = {'equal': 3, 'higher': 2, 'lower': 1, 'unknown': 0}
    top = max(cmps, key=lambda x: rank[x])
    if top == 'equal':
        overall, vres, note = 'PASS', '版本一致', ''
    elif top == 'higher':
        overall, vres, note = 'PASS', '版本偏高', f'实机({",".join(vers)})高于清单({expected["version"]})，按“更高也正确”通过'
    elif top == 'lower':
        overall, vres, note = 'FAIL', '版本过低', f'商店应用版本低于清单：{",".join(vers)} < {expected["version"]}'
    else:
        overall, vres, note = 'WARN', '版本无法解析', ''
    base.update({'actual': ', '.join(vers), 'method': 'STORE_APP(商店应用)', 'ver': vres,
                 'overall': overall, 'note': note})
    return base


def find_matches(expected, collect):
    """在采集数据里找 expected（一条表格记录）对应的实机驱动。
    返回 (matches, method)：matches 是命中的 signed_driver 列表；method 说明匹配方式。"""
    en = norm(expected['name'])
    et = set(tokens(expected['name']))
    signed = collect['signed_drivers']

    # 1) 精确：规范化名完全相等
    m = [d for d in signed if norm(d['DeviceName']) == en]
    if m:
        return m, 'exact精确'
    # 2) 子串：表格名整体是设备名的子串（处理“前缀+后缀型号”）
    m = [d for d in signed if en and en in norm(d['DeviceName'])]
    if m:
        return m, 'substring子串'
    # 3) 词子集：表格名的所有词都出现在设备名里（较模糊）
    if et:
        m = [d for d in signed if et.issubset(set(tokens(d['DeviceName'])))]
        if m:
            return m, 'token词子集(模糊)'
    # 4) 别名正则：在 设备名/厂商 里搜
    for pat in ALIASES.get(expected['name'], []):
        rgx = re.compile(pat, re.I)
        m = [d for d in signed if rgx.search(d['DeviceName'] or '') or rgx.search(d['Provider'] or '')]
        if m:
            return m, f'alias别名(/{pat}/)'
    # 5) 别名回退到 INF 包（把 inf 记录包装成类似结构）
    for pat in ALIASES.get(expected['name'], []):
        rgx = re.compile(pat, re.I)
        m = []
        for p in collect.get('inf_packages', []):
            if rgx.search(p.get('OriginalName', '') or '') or rgx.search(p.get('Provider', '') or '') \
               or rgx.search(p.get('PublishedName', '') or ''):
                m.append({
                    'DeviceName': f"[INF] {p['OriginalName']} ({p['PublishedName']})",
                    'Version': inf_version(p), 'Provider': p['Provider'],
                    'IsSigned': bool(p.get('Signer')), 'Signer': p.get('Signer'),
                    'Class': p.get('Class'), 'DeviceID': p['PublishedName'], '_from_inf': True,
                })
        if m:
            return m, f'alias->INF包(/{pat}/)'
    return [], 'NOT_FOUND'


def uniq_versions(matches):
    """命中的多个子设备可能同版本，去重返回 [(version, is_signed, devicename)]"""
    seen = {}
    for d in matches:
        v = d.get('Version')
        if v not in seen:
            seen[v] = d
    return list(seen.values())


# ------------------------------------------------------------------ 主流程

# 驱动清单固定命名：把清单文件重命名为 driver_test.xlsx 放到本程序同目录即可
EXCEL_FIXED_NAMES = ['driver_test.xlsx', 'driver_test.xlsm', 'driver_test.xls']


def find_excel():
    """查找固定命名的驱动清单 driver_test.xlsx（优先 exe同目录，其次当前工作目录）。"""
    for d in dict.fromkeys([app_dir(), os.getcwd()]):
        for n in EXCEL_FIXED_NAMES:
            f = os.path.join(d, n)
            if os.path.isfile(f):
                return f
    return None


def restart_as_admin():
    """非管理员时，弹UAC以管理员权限重启本程序；返回是否已成功发起提权。
    工作目录指定为 exe 所在目录，保证提权后仍能找到同目录的 driver_test.xlsx 并在此输出报告。"""
    if getattr(sys, 'frozen', False):
        exe = sys.executable                               # 打包后的 exe
        params = subprocess.list2cmdline(sys.argv[1:])
    else:
        exe = sys.executable                               # python.exe
        params = subprocess.list2cmdline([os.path.abspath(sys.argv[0])] + sys.argv[1:])
    try:
        ret = ctypes.windll.shell32.ShellExecuteW(None, 'runas', exe, params, app_dir(), 1)
        return ret > 32
    except Exception:
        return False


def main():
    # 打包成 exe 后，中文系统控制台默认 GBK，遇到 ® 等字符会 UnicodeEncodeError；
    # 统一把 stdout/stderr 重设为 utf-8 且容错，保证不因打印崩溃（报告文件才是最终交付物）。
    for _s in (sys.stdout, sys.stderr):
        try:
            _s.reconfigure(encoding='utf-8', errors='replace')
        except Exception:
            pass

    ap = argparse.ArgumentParser()
    ap.add_argument('--excel', help='驱动清单 xlsx 路径（不填则自动在exe目录/当前目录查找）')
    ap.add_argument('--json', help='目标机采集出的 driver_collect.json（离线模式）')
    ap.add_argument('--local', action='store_true', help='对当前机器现采现比')
    ap.add_argument('--remote', help='远程目标机 IP/主机名（WinRM）')
    ap.add_argument('--user', help='远程用户名')
    ap.add_argument('--report', help='把报告另存为指定文件（.xlsx 或 .txt）')
    args = ap.parse_args()

    # 双击运行(无任何命令行参数)时：结束/出错都停住窗口，方便查看结果
    interactive = (len(sys.argv) == 1)

    def pause_exit(code):
        if interactive:
            try:
                input('\n按回车键退出...')
            except Exception:
                pass
        sys.exit(code)

    here = app_dir()
    ps1 = os.path.join(res_dir(), 'collect_drivers.ps1')

    # 数据来源未指定则默认采集本机(适合把 exe 拷到目标机上双击运行)
    if not (args.json or args.local or args.remote):
        args.local = True
        print('[自动] 未指定数据来源，默认采集【当前机器】进行核对。')

    # --excel 未指定则查找固定命名的清单 driver_test.xlsx（先查清单，缺失即提示，避免无谓提权）
    if not args.excel:
        args.excel = find_excel()
        if not args.excel:
            print('[错误] 未找到驱动清单。请把清单文件重命名为 driver_test.xlsx，'
                  '放到本程序同目录后再运行（或用 --excel 指定路径）。')
            pause_exit(2)
        print(f'[清单] 使用驱动清单: {args.excel}')

    # 权限检查：采集本机(--local)需要管理员权限。非管理员则直接以管理员身份重启自身(弹UAC)。
    # 离线 --json / 远程 --remote 模式仅读文件，不需要提权。
    if args.local and not is_admin():
        print('[权限] 采集本机驱动需要管理员权限，正在以管理员身份重新启动（请在 UAC 弹窗点“是”）...')
        if restart_as_admin():
            sys.exit(0)   # 管理员实例已拉起，退出当前非管理员实例
        print('[权限] 提权未成功（UAC 被取消或失败）。请右键本程序 →「以管理员身份运行」。')
        pause_exit(1)

    try:
        if args.json:
            collect = load_collect(args.json)
        elif args.remote:
            collect = run_collector(ps1, os.path.join(here, 'driver_collect_remote.json'),
                                    remote=args.remote, user=args.user)
        else:  # local
            collect = run_collector(ps1, os.path.join(here, 'driver_collect.json'))
    except Exception as ex:
        print(f'[错误] 采集/读取驱动数据失败: {ex}')
        pause_exit(2)

    try:
        expected_rows = load_excel(args.excel)
    except Exception as ex:
        print(f'[错误] 读取驱动清单失败: {args.excel}\n       {ex}')
        pause_exit(2)

    out = []      # 收集报告文本
    def p(s=''):
        out.append(s)
        print(s)

    meta = collect.get('meta', {})
    p('=' * 78)
    p(f"目标机: {meta.get('ComputerName','?')}  系统: {meta.get('OSCaption','?')} "
      f"Build {meta.get('OSBuild','?')}  采集时间: {meta.get('CollectedAt','?')}  "
      f"管理员采集: {meta.get('IsAdmin','?')}")
    p(f"清单驱动数: {len(expected_rows)}  实机已装驱动: {len(collect['signed_drivers'])}  "
      f"INF包: {len(collect.get('inf_packages',[]))}  异常设备: {len(collect.get('problems',[]))}")
    p('=' * 78)

    fail = 0
    warn = 0
    matched_device_ids = set()   # 记录被表格命中的实机驱动，供检查4排除
    driver_results = []          # 检查1&2&3 逐条结果（供报告）
    unsigned_results = []        # 检查4 未签名列表
    problem_results = []         # 检查5 异常设备列表

    # ===== 检查1/2/3：逐条表格驱动 =====
    p('\n########## 检查1&2&3：表格驱动 存在性 / 版本 / 签名 ##########')
    for e in expected_rows:
        # 商店(UWP)应用类清单项：走 Appx 核对，不参与驱动/INF 匹配
        if e['name'] in STORE_APPS:
            r = check_store_app(e, collect)
            driver_results.append(r)
            if r['overall'] == 'FAIL':
                fail += 1
            elif r['overall'] == 'WARN':
                warn += 1
            tag = {'PASS': '[PASS]', 'FAIL': '[FAIL]', 'WARN': '[WARN]'}[r['overall']]
            p(f"{tag}[商店应用] {e['name']}  期望 {e['version']} / 实机 {r['actual'] or '-'} "
              f"<{r['method']}>")
            if r['note']:
                p(f"        └ {r['note']}")
            continue
        matches, method = find_matches(e, collect)
        if not matches:
            fail += 1
            driver_results.append({
                'name': e['name'], 'vendor': e['vendor'], 'expected': e['version'],
                'actual': '', 'method': 'NOT_FOUND', 'ver': '缺失', 'sign': '',
                'overall': 'FAIL', 'note': '实机未找到该驱动',
            })
            p(f"[FAIL][缺失] {e['name']}  (期望版本 {e['version']}, 厂商 {e['vendor']}) —— 实机未找到该驱动")
            continue

        for d in matches:
            if d.get('DeviceID'):
                matched_device_ids.add(d['DeviceID'])

        vers = uniq_versions(matches)
        # 版本判断：只要有一个命中版本 >= 期望即视为通过（取最优）
        results = [(cmp_ver(d['Version'], e['version']), d) for d in vers]
        best_rank = {'equal': 3, 'higher': 2, 'lower': 1, 'unknown': 0}
        results.sort(key=lambda x: best_rank[x[0]], reverse=True)
        top_cmp, top_dev = results[0]

        vlist = ', '.join(sorted({str(d['Version']) for d in vers}))
        # 版本结论
        overall = 'PASS'
        note = ''
        if top_cmp == 'equal':
            vtag = '[PASS][版本一致]'; vres = '版本一致'
        elif top_cmp == 'higher':
            vtag = '[PASS][版本偏高]'; vres = '版本偏高'; warn += 1
            note = f'实机({top_dev["Version"]})高于清单({e["version"]})，按“更高也正确”通过，建议核对是否更新清单'
        elif top_cmp == 'lower':
            vtag = '[FAIL][版本过低]'; vres = '版本过低'; fail += 1; overall = 'FAIL'
            note = f'实机版本低于清单：{top_dev["DeviceName"]} = {top_dev["Version"]}'
        else:
            vtag = '[WARN][版本无法解析]'; vres = '版本无法解析'; warn += 1
            if overall == 'PASS':
                overall = 'WARN'

        # 签名结论（表内驱动必须签名）
        unsigned = [d for d in matches if not d.get('IsSigned')]
        stag = ''
        sres = '已签名'
        if unsigned:
            stag = '  [FAIL][未签名]'; fail += 1; overall = 'FAIL'; sres = '未签名'

        cnt = f"(命中{len(matches)}个设备)" if len(matches) > 1 else ''
        driver_results.append({
            'name': e['name'], 'vendor': e['vendor'], 'expected': e['version'],
            'actual': vlist + (f'  {cnt}' if cnt else ''), 'method': method,
            'ver': vres, 'sign': sres, 'overall': overall, 'note': note,
        })
        p(f"{vtag} {e['name']}  期望 {e['version']} / 实机 {vlist} {cnt} "
          f"<匹配:{method}>{stag}")
        if note:
            p(f"        └ {note}")

    # ===== 检查4：实机所有驱动签名情况 =====
    p('\n########## 检查4：实机全部驱动签名检查（表外未签名仅WARN） ##########')
    all_unsigned = [d for d in collect['signed_drivers'] if d.get('IsSigned') is False]
    if not all_unsigned:
        p('[PASS] 实机所有已装驱动均已签名。')
    else:
        for d in all_unsigned:
            in_table = d.get('DeviceID') in matched_device_ids
            lvl = 'FAIL' if in_table else 'WARN'
            if in_table:
                fail += 1
            else:
                warn += 1
            unsigned_results.append({
                'level': lvl, 'in_table': in_table, 'name': d['DeviceName'],
                'version': d.get('Version', ''), 'provider': d.get('Provider', ''),
                'class': d.get('Class', ''),
            })
            tag = '表内' if in_table else '表外'
            p(f"[{lvl}][{tag}未签名] {d['DeviceName']}  v{d['Version']}  厂商 {d['Provider']}")

    # ===== 检查5：异常设备 =====
    p('\n########## 检查5：异常设备（黄感/未挂载/无驱动等） ##########')
    problems = collect.get('problems', [])
    if not problems:
        p('[PASS] 未发现在位的异常设备。')
    else:
        for pr in problems:
            warn += 1
            problem_results.append({
                'name': pr.get('Name') or '(无名称)', 'code': pr.get('ErrorCode'),
                'text': pr.get('ErrorText'), 'class': pr.get('Class'), 'id': pr.get('DeviceID'),
            })
            p(f"[WARN][异常] {pr.get('Name') or '(无名称)'}  错误码{pr.get('ErrorCode')}="
              f"{pr.get('ErrorText')}  类:{pr.get('Class')}  {pr.get('DeviceID')}")

    # ===== 汇总 =====
    passed = sum(1 for r in driver_results if r['overall'] == 'PASS')
    p('\n' + '=' * 78)
    p(f"总结：清单 {len(expected_rows)} 条  ->  PASS {passed} | WARN {warn} | FAIL {fail}")
    p('=' * 78)

    summary = {
        'meta': meta, 'expected_total': len(expected_rows),
        'signed_total': len(collect['signed_drivers']),
        'passed': passed, 'warn': warn, 'fail': fail,
        'verdict': 'FAIL' if fail else ('WARN' if warn else 'PASS'),
    }

    # ===== 生成报告 =====
    # 报告路径：--report 指定则用之(扩展名 .txt 出文本，其它出 xlsx)；未指定则自动命名 xlsx
    comp = meta.get('ComputerName', 'device')
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    report_path = args.report or os.path.join(here, f'driver_report_{comp}_{stamp}.xlsx')
    if report_path.lower().endswith('.txt'):
        with open(report_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(out))
    else:
        write_xlsx_report(report_path, summary, driver_results, unsigned_results, problem_results)
    print(f"\n>>> 测试报告已生成: {report_path}")

    pause_exit(1 if fail else 0)


def write_xlsx_report(path, summary, driver_results, unsigned_results, problem_results):
    """生成带颜色分Sheet的Excel测试报告。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    GREEN = PatternFill('solid', fgColor='C6EFCE')   # PASS
    RED = PatternFill('solid', fgColor='FFC7CE')     # FAIL
    ORANGE = PatternFill('solid', fgColor='FFEB9C')  # WARN
    HEAD = PatternFill('solid', fgColor='4472C4')
    head_font = Font(bold=True, color='FFFFFF')
    title_font = Font(bold=True, size=14)
    thin = Side(style='thin', color='D9D9D9')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_of = {'PASS': GREEN, 'FAIL': RED, 'WARN': ORANGE}

    wb = openpyxl.Workbook()

    def style_header(ws, row, ncol):
        for c in range(1, ncol + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = HEAD; cell.font = head_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

    # ---------- Sheet1 汇总 ----------
    ws = wb.active
    ws.title = '汇总'
    m = summary['meta']
    ws['A1'] = '驱动清单核对报告'; ws['A1'].font = title_font
    info = [
        ('目标机', m.get('ComputerName', '?')),
        ('操作系统', f"{m.get('OSCaption','?')} Build {m.get('OSBuild','?')}"),
        ('数据采集时间', m.get('CollectedAt', '?')),
        ('管理员权限采集', m.get('IsAdmin', '?')),
        ('报告生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ('', ''),
        ('清单驱动总数', summary['expected_total']),
        ('实机已装驱动数', summary['signed_total']),
        ('PASS 通过', summary['passed']),
        ('WARN 关注', summary['warn']),
        ('FAIL 不通过', summary['fail']),
        ('总体结论', summary['verdict']),
    ]
    r = 3
    for k, v in info:
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=v)
        if k == '总体结论':
            ws.cell(row=r, column=2).fill = fill_of.get(v, ORANGE)
            ws.cell(row=r, column=2).font = Font(bold=True)
        r += 1
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 55

    # ---------- Sheet2 驱动比对（检查1&2&3） ----------
    ws = wb.create_sheet('驱动比对')
    headers = ['结论', '清单驱动名', '厂商', '期望版本', '实机版本', '版本判断', '签名', '匹配方式', '备注']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for rrow in driver_results:
        ws.append([rrow['overall'], rrow['name'], rrow['vendor'], rrow['expected'],
                   rrow['actual'], rrow['ver'], rrow['sign'], rrow['method'], rrow['note']])
        fill = fill_of.get(rrow['overall'])
        rn = ws.max_row
        for c in range(1, len(headers) + 1):
            ws.cell(row=rn, column=c).border = border
        if fill:
            ws.cell(row=rn, column=1).fill = fill
            ws.cell(row=rn, column=1).font = Font(bold=True)
    widths = [8, 42, 16, 16, 26, 12, 10, 24, 50]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = 'A2'

    # ---------- Sheet3 未签名(检查4) ----------
    ws = wb.create_sheet('未签名驱动')
    headers = ['结论', '范围', '设备名', '版本', '厂商', '类']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    if not unsigned_results:
        ws.append(['PASS', '', '实机所有已装驱动均已签名', '', '', ''])
        ws.cell(row=2, column=1).fill = GREEN
    else:
        for u in unsigned_results:
            ws.append([u['level'], '表内' if u['in_table'] else '表外',
                       u['name'], u['version'], u['provider'], u['class']])
            ws.cell(row=ws.max_row, column=1).fill = fill_of.get(u['level'])
    for i, w in enumerate([8, 8, 45, 18, 28, 16], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ---------- Sheet4 异常设备(检查5) ----------
    ws = wb.create_sheet('异常设备')
    headers = ['结论', '设备名', '错误码', '含义', '类', 'DeviceID']
    ws.append(headers)
    style_header(ws, 1, len(headers))
    if not problem_results:
        ws.append(['PASS', '未发现在位的异常设备', '', '', '', ''])
        ws.cell(row=2, column=1).fill = GREEN
    else:
        for pr in problem_results:
            ws.append(['WARN', pr['name'], pr['code'], pr['text'], pr['class'], pr['id']])
            ws.cell(row=ws.max_row, column=1).fill = ORANGE
    for i, w in enumerate([8, 40, 8, 24, 14, 55], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    wb.save(path)


if __name__ == '__main__':
    main()
