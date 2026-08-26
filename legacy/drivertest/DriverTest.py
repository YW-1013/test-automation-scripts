import wmi
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os
import sys

def get_system_drivers():
    c = wmi.WMI()
    drivers_info = {}

    try:
        drivers = c.Win32_PnPSignedDriver()
        for driver in drivers:
            driver_name = driver.DeviceName
            manufacturer = driver.Manufacturer
            version = driver.DriverVersion
            drivers_info[driver_name] = (manufacturer, version)
    except Exception as e:
        print(f"Error retrieving driver information: {e}")

    return drivers_info

def adjust_column_width(ws):
    for column_cells in ws.columns:
        column_letter = get_column_letter(column_cells[0].column)
        # Check if it's the "检查结果" column
        if column_cells[0].value == "检查结果":
            # Set fixed width for "检查结果" column
            ws.column_dimensions[column_letter].width = 18  # Adjust as needed
        else:
            # Calculate max length for other columns
            max_length = max(len(str(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_letter].width = max_length + 2

def adjust_row_height(ws):
    for row_cells in ws.iter_rows():
        max_length = max(len(str(cell.value)) for cell in row_cells)
        row_number = row_cells[0].row
        ws.row_dimensions[row_number].height = max_length * 0.6 + 10

def compare_drivers(excel_file):
    # 读取Excel表格
    df = pd.read_excel(excel_file)

    # 获取系统驱动信息
    system_drivers = get_system_drivers()

    # 初始化结果列
    df['检查结果'] = ""
    df['实际信息'] = ""

    # 对照驱动信息
    for index, row in df.iterrows():
        driver_name = row['驱动名称'].strip()
        expected_manufacturer = row['驱动厂商'].strip()
        expected_version = row['驱动版本号'].strip()

        if driver_name not in system_drivers:
            df.at[index, '检查结果'] = '不存在'
        else:
            actual_manufacturer, actual_version = system_drivers[driver_name]

            # 检查厂商和版本号是否一致
            if actual_manufacturer != expected_manufacturer and actual_version != expected_version:
                df.at[index, '检查结果'] = '厂商和版本号不一致'
                df.at[index, '实际信息'] = f"{actual_manufacturer}, {actual_version}"
            elif actual_manufacturer != expected_manufacturer:
                df.at[index, '检查结果'] = '厂商不一致'
                df.at[index, '实际信息'] = actual_manufacturer
            elif actual_version != expected_version:
                df.at[index, '检查结果'] = '版本号不一致'
                df.at[index, '实际信息'] = actual_version
            else:
                df.at[index, '检查结果'] = '通过'

    # 写回到原Excel文件
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')  # 假设你的数据在Sheet1中

    # 加载工作簿并调整列宽和行高
    wb = load_workbook(excel_file)
    ws = wb['Sheet1']
    adjust_column_width(ws)
    adjust_row_height(ws)
    wb.save(excel_file)

if __name__ == '__main__':
    current_working_dir = os.path.dirname(os.path.realpath(sys.argv[0]))  # 当前工作目录
    driver_path = os.path.join(current_working_dir,"driver_test.xlsx")
    compare_drivers(driver_path)