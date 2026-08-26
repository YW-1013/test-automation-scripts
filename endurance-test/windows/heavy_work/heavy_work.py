"""
一、多场景负载模拟
1、腾讯双人会议2小时----ok
2、使用高速SSD进行文件压缩和解压操作，模拟混合办公负载，持续30分钟----ok
3、PPT播放，持续30分钟----ok
4、视频格式转换/视频编辑----ok
5、本地4K视频播放，持续30分钟----ok
6、网络压测----ok
"""
import win32con
import wmi
import pyautogui
import subprocess
import os
import time
import psutil
import openpyxl
import pyautogui
import win32gui
import win32api
import sys
import ctypes
import configparser
import shutil
import zipfile
import requests
import json
import cv2
from datetime import datetime
from functools import partial
from openpyxl.styles import Alignment, PatternFill
from moviepy.editor import VideoFileClip
import requests
import json
import psutil
import GPUtil
import wmi
import io
from pymediainfo import MediaInfo
import re

# 窗口状态常量
SW_SHOWMAXIMIZED = 3
SW_SHOWNORMAL = 1
SW_SHOWMINIMIZED = 2

# 初始化配置
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
user32 = ctypes.windll.user32

# 配置文件加载
current_working_dir = os.path.dirname(os.path.realpath(sys.argv[0]))
config = configparser.ConfigParser()
config.read(os.path.join(current_working_dir, 'config.ini'), encoding='utf-8')

# 路径配置
paths = config['Paths']
tenxunmeeting_path = paths.get('tenxunmeeting_path')
urls = config['Urls']
ppt_path = os.path.join(current_working_dir, 'test.pptx')
video_dir = os.path.join(current_working_dir, 'videos')
folder_path = os.path.join(current_working_dir, 'test_folder')
ota_zip_path = os.path.join(current_working_dir, 'H3C_1.0.7.6.zip')
video_path = os.path.join(current_working_dir, 'test.mp4')
# 参数配置
settings = config['Settings']
meeting_code = settings.get('meeting_code')
test_times = settings.getint('test_times')
test_url = urls.get('test_url')

image_dir = os.path.join(current_working_dir, 'images')
join_meeting_path = os.path.join(image_dir, "join_meeting.jpg")
join_meeting_botton = os.path.join(image_dir, "join_meeting_botton.jpg")
exit_meeting_path = os.path.join(image_dir, "exit_meeting.jpg")
log_path = os.path.join(current_working_dir, 'system_status_log.txt')
meeting_time = settings.getint('meeting_time')
ppt_play_time = settings.getint('ppt_play_time')
play_duration = settings.getint('play_duration')


# 全局变量
excel_file = None


# 自动获取视频时长
def get_video_duration(video_path):
    media_info = MediaInfo.parse(video_path)
    for track in media_info.tracks:
        if track.track_type == 'Video':
            time_total = int(float(track.duration))
            return time_total / 1000  # 转换为秒
    raise ValueError("无法获取视频时长")


video_play_time = get_video_duration(video_path)
video_play_time_add = video_play_time + 2

def click_button(image_path, confidence=0.6):
    try:
        location = pyautogui.locateCenterOnScreen(image_path, confidence=confidence)
        if location is not None:
            pyautogui.click(location)
            time.sleep(1)  # 等待界面刷新
        else:
            print(f"未找到按钮: {image_path}")
    except Exception as e:
        print(f"点击按钮时发生错误: {e}")


def create_excel_with_header():
    global excel_file
    now = datetime.now()
    file_name = f"system_test_{now.strftime('%Y%m%d%H%M%S')}.xlsx"
    excel_file = os.path.join(current_working_dir, file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["测试场景", "开始时间", "结束时间", "起始电量(%)", "结束电量(%)",
               "起始电量(mWh)", "结束电量(mWh)", "最大容量(mWh)", "消耗电量(mWh)",
               "剩余电量续航时间预估", "预估总续航时间"]
    ws.append(headers)

    # 设置列宽和样式
    column_widths = {'A': 25, 'B': 18, 'C': 18, 'D': 14, 'E': 14,
                     'F': 16, 'G': 16, 'H': 16, 'I': 16, 'J': 16, 'K': 16}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = CENTER_ALIGNMENT

    wb.save(excel_file)


def get_battery_info():
    battery = psutil.sensors_battery()
    return battery.percent if battery else 0


def write_to_excel(data):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append(data)

    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
    wb.save(excel_file)

def get_battery_capacity_info():
    try:
        c = wmi.WMI(namespace="root\\wmi")
        battery_status = c.BatteryStatus()[0]
        report_path = os.path.join(current_working_dir, "battery-report.html")
        subprocess.run(f'powercfg /batteryreport /output "{report_path}"',
                       shell=True,
                       check=True,
                       creationflags=subprocess.CREATE_NO_WINDOW)
        with open(report_path, 'r', encoding='utf-8') as f:
            content = f.read()

            # 改进版正则表达式
            design_pattern = r'<span class="label">DESIGN CAPACITY</span></td>\s*<td>\s*([\d,]+)\s*mWh'
            full_charge_pattern = r'<span class="label">FULL CHARGE CAPACITY</span></td>\s*<td>\s*([\d,]+)\s*mWh'

            design_match = re.search(design_pattern, content)
            full_match = re.search(full_charge_pattern, content)

            if not design_match or not full_match:
                raise ValueError("无法在报告中找到电池容量信息")
            full_charge = int(full_match.group(1).replace(',', ''))
        return battery_status.RemainingCapacity, full_charge
    except:
        return 0, 0


def record_operation_step(step_name, operation_func, duration=None):
    start_time = datetime.now()
    start_percent = get_battery_info()
    start_cap, design_cap = get_battery_capacity_info()

    if duration:
        operation_func(duration)
    else:
        operation_func()

    end_time = datetime.now()
    delta_seconds = (end_time - start_time).total_seconds()  # 新增时间计算
    end_percent = get_battery_info()
    end_cap, _ = get_battery_capacity_info()

    # 构造带耗时的步骤名称
    formatted_duration = format_duration(delta_seconds)
    step_name_with_duration = f"{step_name}（{formatted_duration}）"

    # 计算续航时间
    delta_seconds = (end_time - start_time).total_seconds()
    consumption = start_cap - end_cap
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate if rate > 0 else 0
        hours, remainder = divmod(remaining, 3600)
        minutes, seconds = divmod(remainder, 60)
        runtime = f"{int(hours)}h{int(minutes)}m{int(seconds)}s"
    else:
        runtime = "N/A"

    # 新增总续航时间计算
    if consumption > 0 and design_cap > 0:
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        hours, remainder = divmod(total_runtime_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        total_runtime = f"{int(hours)}h{int(minutes)}m{int(seconds)}s"
    else:
        total_runtime = "N/A"

    # 写入Excel
    data = [
        step_name,
        start_time.strftime("%m-%d %H:%M:%S"),
        end_time.strftime("%m-%d %H:%M:%S"),
        start_percent,
        end_percent,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    write_to_excel(data)
    return delta_seconds


# 各测试场景功能实现
def tencent_meeting(duration):
    subprocess.Popen(tenxunmeeting_path)
    time.sleep(5)
    pyautogui.hotkey('win', 'up')
    time.sleep(2)
    click_button(join_meeting_path)
    time.sleep(1)
    pyautogui.typewrite(meeting_code, interval=0.1)
    time.sleep(5)
    click_button(join_meeting_botton)
    time.sleep(duration)
    pyautogui.hotkey('alt', 'f4')
    time.sleep(5)
    click_button(exit_meeting_path)
    time.sleep(5)
    pyautogui.hotkey('alt', 'f4')
    time.sleep(2)


def file_operations(folder_path):
    base_folder_name = os.path.basename(folder_path)
    parent_folder = os.path.dirname(folder_path)
    zip_file_name = f"{base_folder_name}_done.zip"
    zip_file_path = os.path.join(parent_folder, zip_file_name)

    with zipfile.ZipFile(zip_file_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                zipf.write(file_path, os.path.relpath(file_path, folder_path))

    return zip_file_path


def decompress_folder(zip_file_path):
    base_file_name = os.path.basename(zip_file_path)
    base_name_without_extension = os.path.splitext(base_file_name)[0].replace('_done', '')
    extract_folder_name = f"{base_name_without_extension}_done"
    parent_folder = os.path.dirname(zip_file_path)
    extract_folder_path = os.path.join(parent_folder, extract_folder_name)

    with zipfile.ZipFile(zip_file_path, 'r') as zipf:
        zipf.extractall(extract_folder_path)

    return extract_folder_path


def cleanup_files(*file_paths):
    for file_path in file_paths:
        if os.path.isfile(file_path):
            os.remove(file_path)
        elif os.path.isdir(file_path):
            shutil.rmtree(file_path)


# 压缩解压文件，压缩解压一次共2分钟50S左右
def zip_and_unzip():
    # Step 1: Compress the folder and rename it
    zipped_file_path = file_operations(folder_path)
    print(f'文件压缩至: {zipped_file_path}')

    # Step 2: Decompress the folder and rename it
    decompressed_folder_path = decompress_folder(zipped_file_path)
    print(f'文件解压至: {decompressed_folder_path}')

    # Step 3: Clean up the files
    cleanup_files(zipped_file_path, decompressed_folder_path)
    print(f'删除文件: {zipped_file_path} and {decompressed_folder_path}')


def is_maximized():
    hwnd = win32gui.GetForegroundWindow()
    if hwnd:
        # 定义 WINDOWPLACEMENT 结构体
        class WINDOWPLACEMENT(ctypes.Structure):
            _fields_ = [("length", ctypes.c_uint),
                        ("flags", ctypes.c_uint),
                        ("showCmd", ctypes.c_uint),
                        ("ptMinPosition", ctypes.wintypes.POINT),
                        ("ptMaxPosition", ctypes.wintypes.POINT),
                        ("rcNormalPosition", ctypes.wintypes.RECT)]

        # 创建一个 WINDOWPLACEMENT 对象并初始化其长度
        placement = WINDOWPLACEMENT()
        placement.length = ctypes.sizeof(WINDOWPLACEMENT)

        # 获取窗口显示状态
        ctypes.windll.user32.GetWindowPlacement(hwnd, ctypes.byref(placement))

        # 检查 showCmd 是否为最大化状态
        if placement.showCmd == SW_SHOWMAXIMIZED:
            return True
    return False


# 全屏操作
def toggle_fullscreen():
    if not is_maximized():
        pyautogui.hotkey('win', 'up')  # 只有在当前不处于全屏时，才尝试进行全屏操


def ppt_playback(duration):
    if not os.path.exists(ppt_path):
        raise FileNotFoundError(f"PPT 文件路径不存在: {ppt_path}")

    # 使用默认程序打开PPT文件（假设WPS Office是默认的PPT处理程序）
    subprocess.Popen(['start', ppt_path], shell=True)

    # 等待PPT文件完全打开
    time.sleep(5)
    toggle_fullscreen()
    time.sleep(5)

    # 模拟按键 'F5' 开始幻灯片放映
    pyautogui.hotkey('f5')

    # 等待放映一段时间（假设30秒）
    time.sleep(duration)

    # 关闭WPS应用（假设WPS已经被激活）
    pyautogui.hotkey('alt', 'f4')
    time.sleep(5)
    pyautogui.hotkey('alt', 'f4')
    time.sleep(5)


def video_conversion():
    # 检查输入文件是否存在
    input_filepath = os.path.join(video_dir, "change_fomat.mp4")
    if not os.path.exists(input_filepath):
        print(f"Input file does not exist: {input_filepath}")
        return

    # 确定输出文件路径
    output_filepath = os.path.join(video_dir, "change_fomat.mkv")
    print(output_filepath)

    # 使用moviepy进行格式转换
    try:
        clip = VideoFileClip(input_filepath)
        clip.write_videofile(output_filepath, codec='libx264')
        print(f"Converted {video_dir} to mkv")
    except Exception as e:
        print(f"Failed to convert {video_dir}: {e}")

    time.sleep(20)

    os.remove(output_filepath)

def video_operation(duration):
    # 记录初始测试开始时间
    test_start_time = time.time()

    # 打开视频并等待播放器启动
    os.startfile(video_path)
    time.sleep(5)  # 关键修改：等待播放器完全启动

    # 记录视频循环开始时间
    video_start_time = time.time()  # 关键修改：从实际播放开始计时

    while (time.time() - test_start_time) < duration:
        current_time = time.time()
        elapsed = current_time - video_start_time

        # 总测试时间检查
        if (current_time - test_start_time) >= duration:
            break

        # 视频播放完成检查（增加1秒容错）
        if elapsed >= video_play_time_add:
            # 执行暂停-全屏切换操作
            pyautogui.press('space')  # 暂停
            time.sleep(0.8)  # 增加延迟确保生效
            pyautogui.hotkey('alt', 'enter')  # 全屏

            # 重置视频计时（以当前操作为新起点）
            video_start_time = time.time()  # 关键修改：完全重置时间基准

        time.sleep(0.5)

    # 关闭处理
    pyautogui.hotkey('alt', 'f4')
    time.sleep(2)


def download_file_and_delete_afterwards():
    """
    下载文件到指定目录，下载完成后删除文件。

    Parameters:
    url (str): 文件下载地址
    download_dir (str): 文件下载保存的目录
    """
    # 确保下载目录存在
    if not os.path.exists(current_working_dir):
        os.makedirs(current_working_dir)

    # 确定文件名
    filename = test_url.split("/")[-1]
    file_path = os.path.join(current_working_dir, filename)

    try:
        # 发送 HTTP 请求以获取文件
        response = requests.get(test_url, stream=True)
        response.raise_for_status()  # 检查请求是否成功

        # 打开文件以写入下载数据
        with open(file_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:  # 过滤出空块
                    f.write(chunk)

        print(f"文件下载完成，保存路径: {file_path}")

    except requests.RequestException as e:
        print(f"下载过程中发生错误: {e}")
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"已删除下载失败的文件: {file_path}")
        return

    try:
        # 下载完成后删除文件
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"文件已删除: {file_path}")
        else:
            print(f"文件未找到: {file_path}")
    except Exception as e:
        print(f"删除文件时发生错误: {e}")


def get_token():
    url = "http://your-server.example.com/ota-api/auth/login"

    payload = json.dumps({
        "username": "admin",
        "password": "YOUR_PASSWORD"
    })
    headers = {
        'User-Agent': 'Apifox/1.0.0 (https://apifox.com)',
        'Content-Type': 'application/json'
    }

    response = requests.request("POST", url, headers=headers, data=payload)

    response_data = response.json()
    access_token = response_data.get('data', {}).get('access_token')
    return access_token


def upload_files(token):
    url = "http://your-server.example.com/ota/file/full-package/upload-package"

    payload = {}
    files = [
        ('file', (ota_zip_path, open(ota_zip_path, 'rb'), 'application/zip'))
    ]
    headers = {
        'Authorization': token,
        'User-Agent': 'Apifox/1.0.0 (https://apifox.com)'
    }

    response = requests.request("POST", url, headers=headers, data=payload, files=files)
    return response.text


# 进行网络压测
def network_stress():
    download_file_and_delete_afterwards()
    time.sleep(5)
    upload_files(get_token())


def format_duration(seconds):
    if seconds < 60:
        return f"{int(seconds)}s"
    elif seconds < 3600:
        return f"{int(seconds//60)}min{int(seconds%60)}s"
    else:
        return f"{int(seconds//3600)}h{int((seconds%3600)//60)}min"


def add_cycle_summary(cycle_num, first_start_row, current_end_row, use_run_time):
    """根据总起始行和当前结束行生成累计循环汇总"""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 新增绿色填充样式
    highlight_fill = PatternFill(start_color='C6EFCE',  # 浅绿色背景
                                 end_color='C6EFCE',
                                 fill_type='solid')

    # 获取总起始行数据（第一轮的第一条记录）
    first_row = ws[first_start_row]
    # 获取当前轮的最后一条记录
    last_row = ws[current_end_row]

    # 提取时间数据（需处理无年份的日期）
    start_time_str = f"2025-{first_row[1].value}"  # 假设测试年份为2025
    end_time_str = f"2025-{last_row[2].value}"
    start_time = datetime.strptime(start_time_str, "%Y-%m-%d %H:%M:%S")
    end_time = datetime.strptime(end_time_str, "%Y-%m-%d %H:%M:%S")

    # 计算时间差
    delta_seconds = (end_time - start_time).total_seconds()

    # 提取电量数据
    start_cap = first_row[5].value
    end_cap = last_row[6].value
    design_cap = first_row[7].value
    consumption = start_cap - end_cap

    # 计算续航时间（逻辑与原函数一致）
    runtime, total_runtime = "N/A", "N/A"
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate
        runtime = format_runtime(remaining)
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        total_runtime = format_runtime(total_runtime_seconds)

    # 构建汇总行数据
    summary_data = [
        f"{cycle_num}轮循环（{use_run_time}）",
        first_row[1].value,
        last_row[2].value,
        first_row[3].value,
        last_row[4].value,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    # 写入并保存
    ws.append(summary_data)
    # 应用居中样式
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = highlight_fill  # 添加背景色
    wb.save(excel_file)


def main():
    create_excel_with_header()
    first_cycle_start_row = None
    for cycle in range(test_times):
        try:
            # 获取当前起始行
            wb = openpyxl.load_workbook(excel_file)
            start_row = wb.active.max_row
            wb.close()

            if first_cycle_start_row is None:
                first_cycle_start_row = start_row + 1
            total_duration = 0  # 新增总耗时统计
            # 测试场景配置（使用partial绑定参数）
            test_scenarios = [
                (f"腾讯会议", partial(tencent_meeting, duration=meeting_time)),
                ("文件压缩解压操作", partial(zip_and_unzip)),
                (f"PPT播放", partial(ppt_playback, duration=ppt_play_time)),
                ("视频格式转换", video_conversion),
                (f"本地4K播放", partial(video_operation, duration=play_duration)),
                ("网络压力测试", network_stress)
            ]

            # 执行测试步骤并收集耗时
            for step_name, func in test_scenarios:
                delta = record_operation_step(step_name, func)
                total_duration += delta

            # 添加当前轮次汇总
            wb = openpyxl.load_workbook(excel_file)
            current_end_row = wb.active.max_row
            wb.close()

            add_cycle_summary(
                cycle_num=cycle + 1,
                first_start_row=first_cycle_start_row,
                current_end_row=current_end_row,
                use_run_time=format_duration(total_duration)  # 使用真实耗时
            )
        except Exception as e:
            print(f"第{cycle + 1}轮测试失败: {str(e)}")


if __name__ == "__main__":
    main()
