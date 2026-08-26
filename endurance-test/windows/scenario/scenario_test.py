import win32con
import wmi
import subprocess
import os
import time
import psutil
import openpyxl
import pyautogui
import win32gui
import win32api
import sys
import ctypes
import configparser
from docx import Document
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from functools import partial
import win32com.client
from openpyxl.styles import Alignment
import io
from openpyxl.styles import PatternFill
import re
from pymediainfo import MediaInfo
import uiautomator2 as u2


sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
# 窗口状态常量
SW_SHOWMAXIMIZED = 3

# 定义需要的常量和函数
user32 = ctypes.windll.user32

# 配置加载
current_working_dir = os.path.dirname(os.path.realpath(sys.argv[0]))
config = configparser.ConfigParser()
config.read(os.path.join(current_working_dir, 'config.ini'), encoding='utf-8')

# 路径配置
paths =config['Paths']
weixin_path = paths.get("weixin_path")
jianying_path = paths.get("jianying_path")
docx_path = os.path.join(current_working_dir, 'test.docx')
edge_driver = os.path.join(current_working_dir, 'edgedriver_win64', 'msedgedriver.exe')

#点击坐标配置
Locations = config['Locations']
weixin_click1x = Locations.getint('weixin_click1x')
weixin_click1y = Locations.getint('weixin_click1y')
weixin_click2x = Locations.getint('weixin_click2x')
weixin_click2y = Locations.getint('weixin_click2y')
weixin_click3x = Locations.getint('weixin_click3x')
weixin_click3y = Locations.getint('weixin_click3y')
weixin_phone_clickx = Locations.getint('weixin_phone_clickx')
weixin_phone_clicky = Locations.getint('weixin_phone_clicky')
jianying_click1x = Locations.getint('jianying_click1x')
jianying_click1y = Locations.getint('jianying_click1y')
jianying_click2x = Locations.getint('jianying_click2x')
jianying_click2y = Locations.getint('jianying_click2y')
jianying_click3x = Locations.getint('jianying_click3x')
jianying_click3y = Locations.getint('jianying_click3y')

# 参数配置
settings = config['Settings']
word_time = settings.getint('word_time')
web_time = settings.getint('web_time')
weixin_time = settings.getint('weixin_time')
video_time = settings.getint('video_time')
jianying_time = settings.getint('jianying_time')
test_times = settings.getint('test_times')



# 全局变量
excel_file = None


# 自动获取视频时长
def get_video_duration(video_path):
    media_info = MediaInfo.parse(video_path)
    for track in media_info.tracks:
        if track.track_type == 'Video':
            time_total = int(float(track.duration))
            return time_total / 1000  # 转换为秒
    raise ValueError("无法获取视频时长")


video_path = os.path.join(current_working_dir,'test.mkv')
video_play_time = get_video_duration(video_path)
video_play_time_add = video_play_time + 2


def set_english_keyboard():
    # 美国英语键盘布局代码
    ENG_KEYBOARD = 0x04090409
    hwnd = win32gui.GetForegroundWindow()
    win32api.LoadKeyboardLayout('00000409', 1)  # 加载美式键盘
    win32gui.SendMessage(hwnd, win32con.WM_INPUTLANGCHANGEREQUEST, 0, ENG_KEYBOARD)


def create_excel_with_header():
    global excel_file
    now = datetime.now()
    file_name = f"win_result_{now.strftime('%Y%m%d%H%M%S')}.xlsx"
    excel_file = os.path.join(current_working_dir, file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["测试场景", "起始时间", "结束时间", "起始电量（%）", "结束电量（%）",
               "起始电量（mWh）", "结束电量（mWh）", "最大容量（mWh）", "消耗电量（mWh）", "剩余电量续航时间预估", "预估总续航时间"]
    ws.append(headers)

    # 列宽配置字典（单位：字符）
    column_widths = {
        "A": 30,  # 操作步骤
        "B": 15,  # 起始时间
        "C": 15,  # 结束时间
        "D": 14,  # 起始电量%
        "E": 14,  # 结束电量%
        "F": 15,  # 起始容量
        "G": 15,  # 结束容量
        "H": 15,  # 设计容量
        "I": 15,  # 消耗电量
        "J": 15,  # 剩余电量续航时间
        "K": 15  # 预估总续航
    }

    # 应用列宽
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置居中样式
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = CENTER_ALIGNMENT

    wb.save(excel_file)
    return excel_file


def is_maximized():
    hwnd = win32gui.GetForegroundWindow()
    if hwnd:
        # 定义 WINDOWPLACEMENT 结构体
        class WINDOWPLACEMENT(ctypes.Structure):
            _fields_ = [("length", ctypes.c_uint),
                        ("flags", ctypes.c_uint),
                        ("showCmd", ctypes.c_uint),
                        ("ptMinPosition", ctypes.wintypes.POINT),
                        ("ptMaxPosition", ctypes.wintypes.POINT),
                        ("rcNormalPosition", ctypes.wintypes.RECT)]

        # 创建一个 WINDOWPLACEMENT 对象并初始化其长度
        placement = WINDOWPLACEMENT()
        placement.length = ctypes.sizeof(WINDOWPLACEMENT)

        # 获取窗口显示状态
        ctypes.windll.user32.GetWindowPlacement(hwnd, ctypes.byref(placement))

        # 检查 showCmd 是否为最大化状态
        if placement.showCmd == SW_SHOWMAXIMIZED:
            return True
    return False


#全屏操作
def toggle_fullscreen():
    if not is_maximized():
        pyautogui.hotkey('win', 'up')  # 只有在当前不处于全屏时，才尝试进行全屏操作

# 最小化所有窗口
def minimize_all_windows():
    # 发送 Win + D 快捷键两次，可以最小化然后还原所有窗口
    user32.keybd_event(0x5B, 0, 0, 0)  # Press Left Windows
    user32.keybd_event(0x44, 0, 0, 0)  # Press D
    user32.keybd_event(0x44, 0, 2, 0)  # Release D
    user32.keybd_event(0x5B, 0, 2, 0)  # Release Left Windows


def write_to_excel(data):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append(data)

    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
    wb.save(excel_file)


def get_battery_info():
    battery = psutil.sensors_battery()
    return battery.percent if battery else 0


def get_battery_capacity_info():
    try:
        c = wmi.WMI(namespace="root\\wmi")
        battery_status = c.BatteryStatus()[0]
        report_path = os.path.join(current_working_dir, "battery-report.html")
        subprocess.run(f'powercfg /batteryreport /output "{report_path}"',
                       shell=True,
                       check=True,
                       creationflags=subprocess.CREATE_NO_WINDOW)
        with open(report_path, 'r', encoding='utf-8') as f:
            content = f.read()

            # 改进版正则表达式
            design_pattern = r'<span class="label">DESIGN CAPACITY</span></td>\s*<td>\s*([\d,]+)\s*mWh'
            full_charge_pattern = r'<span class="label">FULL CHARGE CAPACITY</span></td>\s*<td>\s*([\d,]+)\s*mWh'

            design_match = re.search(design_pattern, content)
            full_match = re.search(full_charge_pattern, content)

            if not design_match or not full_match:
                raise ValueError("无法在报告中找到电池容量信息")
            full_charge = int(full_match.group(1).replace(',', ''))
        return battery_status.RemainingCapacity, full_charge
    except:
        return 0, 0


def record_operation_step(step_name, operation_func, duration=None):
    """记录操作步骤的核心函数"""
    # 记录开始数据
    start_time = datetime.now()
    start_percent = get_battery_info()
    start_cap, design_cap = get_battery_capacity_info()

    # 执行实际操作
    if duration:
        operation_func(duration)
    else:
        operation_func()

    # 记录结束数据
    end_time = datetime.now()
    end_percent = get_battery_info()
    end_cap, _ = get_battery_capacity_info()

    # 计算指标
    delta_seconds = (end_time - start_time).total_seconds()
    consumption = start_cap - end_cap
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate if rate > 0 else 0
        hours, remainder = divmod(remaining, 3600)
        minutes, seconds = divmod(remainder, 60)
        runtime = f"{int(hours)}h{int(minutes)}m{int(seconds)}s"
    else:
        runtime = "N/A"

    # 新增总续航时间计算
    if consumption > 0 and design_cap > 0:
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        hours, remainder = divmod(total_runtime_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        total_runtime = f"{int(hours)}h{int(minutes)}m{int(seconds)}s"
    else:
        total_runtime = "N/A"


    # 写入Excel
    data = [
        step_name,
        start_time.strftime("%m-%d %H:%M:%S"),
        end_time.strftime("%m-%d %H:%M:%S"),
        start_percent,
        end_percent,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    write_to_excel(data)


# 以下是各功能操作的实现（需要保持原有功能不变，只修改调用方式）
def word_operation(duration):
    """Word文档操作"""
    if os.path.exists(docx_path):
        os.remove(docx_path)
    Document().save(docx_path)
    time.sleep(1)
    os.startfile(docx_path)
    time.sleep(3)

    toggle_fullscreen()
    time.sleep(2)
    set_english_keyboard()
    start_time = time.time()

    while time.time() - start_time < duration:
        pyautogui.write("To be, or not to be - that is the question.", interval=0.1)

    pyautogui.hotkey('ctrl', 's')
    time.sleep(1)
    pyautogui.hotkey('alt', 'f4')
    time.sleep(2)
    processes = ['WINWORD.EXE', 'wps.exe']
    for p in processes:
        subprocess.run(f'taskkill /F /IM {p}', shell=True)


def web_test(duration):
    service = Service(executable_path=edge_driver)
    options = webdriver.EdgeOptions()
    options.add_argument("--start-maximized")  # 强制初始最大化

    with webdriver.Edge(service=service, options=options) as driver:
        driver.get("https://www.bilibili.com")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
        swipe_times = 0
        start_time = time.time()
        while (time.time() - start_time) < duration:
            # 滚动页面并等待
            driver.execute_script("window.scrollBy(0, 500);")
            swipe_times += 1
            time.sleep(3)
            if swipe_times % 100 == 0:
                pyautogui.hotkey('Home')
            # 强制检查时间以防最后一次等待超时
            if (time.time() - start_time) >= duration:
                break

def weixin_chat(duration):
    # 静默模式参数配置
    silent_args = {
        'stdout': subprocess.DEVNULL,  # 禁止标准输出
        'stderr': subprocess.DEVNULL,  # 禁止错误输出
        'shell': True
    }
    subprocess.Popen(weixin_path, **silent_args)  # 打开微信
    time.sleep(5)
    toggle_fullscreen()  # 微信全屏
    time.sleep(5)
    pyautogui.click(x=weixin_click1x, y=weixin_click1y)#点击指定人员头像
    time.sleep(5)
    pyautogui.click(x=weixin_click2x, y=weixin_click2y)#点击视频聊天
    d = u2.connect()
    time.sleep(5)
    d.click(weixin_phone_clickx,weixin_phone_clicky)#手机上点击接电话
    time.sleep(duration)
    pyautogui.click(x=weixin_click3x, y=weixin_click3y)#点击结束聊天
    time.sleep(2)


def video_operation(duration):
    # 记录初始测试开始时间
    test_start_time = time.time()

    # 打开视频并等待播放器启动
    os.startfile(video_path)
    time.sleep(5)  # 关键修改：等待播放器完全启动

    # 记录视频循环开始时间
    video_start_time = time.time()  # 关键修改：从实际播放开始计时

    while (time.time() - test_start_time) < duration:
        current_time = time.time()
        elapsed = current_time - video_start_time

        # 总测试时间检查
        if (current_time - test_start_time) >= duration:
            break

        # 视频播放完成检查（增加1秒容错）
        if elapsed >= video_play_time_add:

            # 执行暂停-全屏切换操作
            pyautogui.press('space')  # 暂停
            time.sleep(0.8)  # 增加延迟确保生效
            pyautogui.hotkey('alt', 'enter')  # 全屏

            # 重置视频计时（以当前操作为新起点）
            video_start_time = time.time()  # 关键修改：完全重置时间基准

        time.sleep(0.5)

    # 关闭处理
    pyautogui.hotkey('alt', 'f4')
    time.sleep(2)

def jianying_operation(duration):
    processes = ['JianyingPro.exe']
    # 静默模式参数配置
    silent_args = {
        'stdout': subprocess.DEVNULL,  # 禁止标准输出
        'stderr': subprocess.DEVNULL,  # 禁止错误输出
        'shell': True
    }
    subprocess.Popen(jianying_path, **silent_args)#打开剪映
    time.sleep(40)
    pyautogui.click(x=jianying_click1x, y=jianying_click1y)#点击草稿
    time.sleep(30)
    pyautogui.click(x=jianying_click2x, y=jianying_click2y)#点击导出
    time.sleep(5)
    pyautogui.click(x=jianying_click3x, y=jianying_click3y)#点击确定导出
    time.sleep(duration)
    for p in processes:
        check = subprocess.run(f'tasklist /FI "IMAGENAME eq {p}"', **silent_args)
        if check.returncode == 0:  # returncode=0表示找到进程
            # 静默终止进程
            subprocess.run(f'taskkill /F /IM {p}', **silent_args)

def format_duration(seconds):
    """秒转分钟格式化（保留两位小数）"""
    minutes = seconds / 60
    if minutes.is_integer():
        return f"{int(minutes)}min"
    return f"{minutes:.2f}min".rstrip('0').rstrip('.') # 智能去除多余零

def calculate_runtime(start_time_str, end_time_str, consumption, end_cap):
    """计算续航时间预估"""
    start_time = datetime.strptime(start_time_str, "%Y-%m-%d %H:%M:%S")
    end_time = datetime.strptime(end_time_str, "%Y-%m-%d %H:%M:%S")
    delta_seconds = (end_time - start_time).total_seconds()

    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate if rate > 0 else 0
        hours, remainder = divmod(remaining, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{int(hours)}h{int(minutes)}m{int(seconds)}s"
    return "N/A"


def format_runtime(seconds):
    """统一格式化时间"""
    hours, remainder = divmod(seconds, 3600)
    minutes, sec = divmod(remainder, 60)
    return f"{int(hours)}h{int(minutes)}m{int(sec)}s"


def add_cycle_summary(cycle_num, first_start_row, current_end_row,use_run_time):
    """根据总起始行和当前结束行生成累计循环汇总"""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 新增绿色填充样式
    highlight_fill = PatternFill(start_color='C6EFCE',  # 浅绿色背景
                               end_color='C6EFCE',
                               fill_type='solid')

    # 获取总起始行数据（第一轮的第一条记录）
    first_row = ws[first_start_row]
    # 获取当前轮的最后一条记录
    last_row = ws[current_end_row]

    # 提取时间数据（需处理无年份的日期）
    start_time_str = f"2025-{first_row[1].value}"  # 假设测试年份为2025
    end_time_str = f"2025-{last_row[2].value}"
    start_time = datetime.strptime(start_time_str, "%Y-%m-%d %H:%M:%S")
    end_time = datetime.strptime(end_time_str, "%Y-%m-%d %H:%M:%S")

    # 计算时间差
    delta_seconds = (end_time - start_time).total_seconds()

    # 提取电量数据
    start_cap = first_row[5].value
    end_cap = last_row[6].value
    design_cap = first_row[7].value
    consumption = start_cap - end_cap

    # 计算续航时间（逻辑与原函数一致）
    runtime, total_runtime = "N/A", "N/A"
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate
        runtime = format_runtime(remaining)
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        total_runtime = format_runtime(total_runtime_seconds)

    # 构建汇总行数据
    summary_data = [
        f"{cycle_num}轮循环（{use_run_time}）",
        first_row[1].value,
        last_row[2].value,
        first_row[3].value,
        last_row[4].value,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    # 写入并保存
    ws.append(summary_data)
    # 应用居中样式
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = highlight_fill  # 添加背景色
    wb.save(excel_file)

def main():
    set_english_keyboard()
    create_excel_with_header()

    # 从配置读取各应用独立时间参数
    app_durations = {
        'word': word_time,
        'web': web_time,
        'weixin': weixin_time,
        'video': video_time,
        'jianying': jianying_time
    }

    first_cycle_start_row = None  # 记录总起始行

    for cycle in range(test_times):
        # 获取当前Excel行号
        wb = openpyxl.load_workbook(excel_file)
        start_row = wb.active.max_row
        wb.close()

        if first_cycle_start_row is None:
            first_cycle_start_row = start_row + 1

        # 构建测试步骤（带独立时间参数）
        test_steps = [
            (f"Word文档编辑（{format_duration(app_durations['word'])}）",
             partial(word_operation, duration=app_durations['word'])),
            (f"网页浏览（{format_duration(app_durations['web'])}）",
             partial(web_test, duration=app_durations['web'])),
            (f"微信视频聊天（{format_duration(app_durations['weixin'])}）",
             partial(weixin_chat, duration=app_durations['weixin'])),
            (f"4K视频播放（{format_duration(app_durations['video'])}）",
             partial(video_operation, duration=app_durations['video'])),
            (f"剪映导出（{format_duration(app_durations['jianying'])}）",
             partial(jianying_operation, duration=app_durations['jianying']))
        ]

        # 执行测试步骤
        for step_name, func in test_steps:
            record_operation_step(step_name, func)

        # 计算单轮总耗时
        cycle_total_time = sum(app_durations.values())
        time_label_cycle = format_duration(cycle_total_time)

        # 添加循环汇总
        wb = openpyxl.load_workbook(excel_file)
        current_end_row = wb.active.max_row
        wb.close()

        add_cycle_summary(
            cycle_num=cycle + 1,
            first_start_row=first_cycle_start_row,
            current_end_row=current_end_row,
            use_run_time=time_label_cycle
        )

if __name__ == '__main__':
    main()