import win32con
import wmi
import subprocess
import os
import time
import psutil
import openpyxl
import pyautogui
import win32gui
import win32api
import sys
import ctypes
import configparser
from docx import Document
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from functools import partial
import win32com.client
from openpyxl.styles import Alignment
import io
from openpyxl.styles import PatternFill
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
# 窗口状态常量
SW_SHOWMAXIMIZED = 3
SW_SCROLL = 0x100
# 定义需要的常量和函数
user32 = ctypes.windll.user32

# 配置加载
current_working_dir = os.path.dirname(os.path.realpath(sys.argv[0]))
config = configparser.ConfigParser()
config.read(os.path.join(current_working_dir, 'config.ini'), encoding='utf-8')

# 路径配置
paths = config['Paths']
wangyiyun_path = paths.get('wangyiyun_path')
qq_path = paths.get('qq_path')
docx_path = os.path.join(current_working_dir, 'test.docx')
xlsx_path = os.path.join(current_working_dir, 'test.xlsx')
ppt_path = os.path.join(current_working_dir, 'test.pptx')
cpu_z_path = os.path.join(current_working_dir, 'cpuz.exe')
video_url = config['Urls'].get('video_url')
edge_driver = os.path.join(current_working_dir, 'edgedriver_win64', 'msedgedriver.exe')
# 获取Edge用户数据目录路径
user_data_dir = os.path.join(
    os.environ['LOCALAPPDATA'],
    'Microsoft', 'Edge', 'UserData'
)

# 参数配置
settings = config['Settings']
test_use_times = settings.getint('test_use_times')
cpu_z_times = settings.getint('cpu_z_times')
test_times = settings.getint('test_times')
DEBUG_TYPE = settings.get('debug')
# 全局变量
excel_file = None

def idle_operation(duration):
    """静止等待操作"""
    time.sleep(duration)

def set_english_keyboard():
    # 美国英语键盘布局代码
    ENG_KEYBOARD = 0x04090409
    hwnd = win32gui.GetForegroundWindow()
    win32api.LoadKeyboardLayout('00000409', 1)  # 加载美式键盘
    win32gui.SendMessage(hwnd, win32con.WM_INPUTLANGCHANGEREQUEST, 0, ENG_KEYBOARD)


def create_excel_with_header():
    global excel_file
    now = datetime.now()
    file_name = f"win_result_{now.strftime('%Y%m%d%H%M%S')}.xlsx"
    excel_file = os.path.join(current_working_dir, file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["操作步骤", "起始时间", "结束时间", "起始电量（%）", "结束电量（%）",
               "起始电量（mWh）", "结束电量（mWh）", "最大容量（mWh）", "消耗电量（mWh）", "剩余电量续航时间预估", "预估总续航时间"]
    ws.append(headers)

    # 列宽配置字典（单位：字符）
    column_widths = {
        "A": 30,  # 操作步骤
        "B": 15,  # 起始时间
        "C": 15,  # 结束时间
        "D": 14,  # 起始电量%
        "E": 14,  # 结束电量%
        "F": 15,  # 起始容量
        "G": 15,  # 结束容量
        "H": 15,  # 设计容量
        "I": 15,  # 消耗电量
        "J": 15,  # 可续航时间
        "K": 15  # 预估总续航
    }

    # 应用列宽
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置居中样式
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = CENTER_ALIGNMENT

    wb.save(excel_file)
    return excel_file


def is_maximized():
    hwnd = win32gui.GetForegroundWindow()
    if hwnd:
        # 定义 WINDOWPLACEMENT 结构体
        class WINDOWPLACEMENT(ctypes.Structure):
            _fields_ = [("length", ctypes.c_uint),
                        ("flags", ctypes.c_uint),
                        ("showCmd", ctypes.c_uint),
                        ("ptMinPosition", ctypes.wintypes.POINT),
                        ("ptMaxPosition", ctypes.wintypes.POINT),
                        ("rcNormalPosition", ctypes.wintypes.RECT)]

        # 创建一个 WINDOWPLACEMENT 对象并初始化其长度
        placement = WINDOWPLACEMENT()
        placement.length = ctypes.sizeof(WINDOWPLACEMENT)

        # 获取窗口显示状态
        ctypes.windll.user32.GetWindowPlacement(hwnd, ctypes.byref(placement))

        # 检查 showCmd 是否为最大化状态
        if placement.showCmd == SW_SHOWMAXIMIZED:
            return True
    return False


#全屏操作
def toggle_fullscreen():
    if not is_maximized():
        pyautogui.hotkey('win', 'up')  # 只有在当前不处于全屏时，才尝试进行全屏操作

# 最小化所有窗口
def minimize_all_windows():
    # 发送 Win + D 快捷键两次，可以最小化然后还原所有窗口
    user32.keybd_event(0x5B, 0, 0, 0)  # Press Left Windows
    user32.keybd_event(0x44, 0, 0, 0)  # Press D
    user32.keybd_event(0x44, 0, 2, 0)  # Release D
    user32.keybd_event(0x5B, 0, 2, 0)  # Release Left Windows


def write_to_excel(data):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append(data)

    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
    wb.save(excel_file)


def get_battery_info():
    battery = psutil.sensors_battery()
    return battery.percent if battery else 0


def get_battery_capacity_info():
    try:
        c = wmi.WMI(namespace="root\\wmi")
        battery_status = c.BatteryStatus()[0]
        report_path = os.path.join(current_working_dir, "battery-report.html")
        subprocess.run(f'powercfg /batteryreport /output "{report_path}"',
                       shell=True,
                       check=True,
                       creationflags=subprocess.CREATE_NO_WINDOW)
        with open(report_path, 'r', encoding='utf-8') as f:
            content = f.read()

            # 改进版正则表达式
            design_pattern = r'<span class="label">DESIGN CAPACITY</span></td>\s*<td>\s*([\d,]+)\s*mWh'
            full_charge_pattern = r'<span class="label">FULL CHARGE CAPACITY</span></td>\s*<td>\s*([\d,]+)\s*mWh'

            design_match = re.search(design_pattern, content)
            full_match = re.search(full_charge_pattern, content)

            if not design_match or not full_match:
                raise ValueError("无法在报告中找到电池容量信息")
            full_charge = int(full_match.group(1).replace(',', ''))
        return battery_status.RemainingCapacity, full_charge
    except:
        return 0, 0


def record_operation_step(step_name, operation_func, duration=None):
    """记录操作步骤的核心函数"""
    # 记录开始数据
    start_time = datetime.now()
    start_percent = get_battery_info()
    start_cap, design_cap = get_battery_capacity_info()

    # 执行实际操作
    if duration:
        operation_func(duration)
    else:
        operation_func()

    # 记录结束数据
    end_time = datetime.now()
    end_percent = get_battery_info()
    end_cap, _ = get_battery_capacity_info()

    # 计算指标
    delta_seconds = (end_time - start_time).total_seconds()
    consumption = start_cap - end_cap
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate if rate > 0 else 0
        hours, remainder = divmod(remaining, 3600)
        minutes, seconds = divmod(remainder, 60)
        runtime = f"{int(hours)}h{int(minutes)}m{int(seconds)}s"
    else:
        runtime = "N/A"

    # 新增总续航时间计算
    if consumption > 0 and design_cap > 0:
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        hours, remainder = divmod(total_runtime_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        total_runtime = f"{int(hours)}h{int(minutes)}m{int(seconds)}s"
    else:
        total_runtime = "N/A"


    # 写入Excel
    data = [
        step_name,
        start_time.strftime("%m-%d %H:%M:%S"),
        end_time.strftime("%m-%d %H:%M:%S"),
        start_percent,
        end_percent,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    write_to_excel(data)


def is_page_bottom(driver):
    """判断是否滚动到页面底部"""
    try:
        scroll_height = driver.execute_script("return document.body.scrollHeight")
        scroll_top = driver.execute_script("return document.documentElement.scrollTop || window.pageYOffset")
        window_height = driver.execute_script("return window.innerHeight")
        return (scroll_top + window_height) >= scroll_height
    except:
        return False



# 以下是各功能操作的实现（需要保持原有功能不变，只修改调用方式）
def backend_preparation(duration):
    """后台准备"""
    processes = ['cloudmusic.exe', 'QQ.exe', 'EXCEL.EXE', 'POWERPNT.EXE', 'WINWORD.EXE']
    # 静默模式参数配置
    silent_args = {
        'stdout': subprocess.DEVNULL,  # 禁止标准输出
        'stderr': subprocess.DEVNULL,  # 禁止错误输出
        'shell': True
    }
    for p in processes:
        check = subprocess.run(f'tasklist /FI "IMAGENAME eq {p}"', **silent_args)
        if check.returncode == 0:  # returncode=0表示找到进程
            # 静默终止进程
            subprocess.run(f'taskkill /F /IM {p}', **silent_args)
    subprocess.Popen(qq_path, **silent_args)
    time.sleep(15)
    subprocess.Popen(wangyiyun_path, **silent_args)
    #等待网易云完全启动
    time.sleep(15)
    # 模拟按下F5键,
    pyautogui.press('f5')
    time.sleep(duration)


def cpu_z(duration):
    pyautogui.hotkey('win', 'm')
    time.sleep(2)
    os.startfile(cpu_z_path)
    time.sleep(10)
    pyautogui.click(1734, 552)
    time.sleep(2)
    start_time = datetime.now()
    while (datetime.now() - start_time).total_seconds() < duration:
        # 模拟向下滑动（屏幕Y轴80%位置滑动到20%位置）
        pyautogui.click(1154, 1048)
        time.sleep(15)  # 刷新间隔
    pyautogui.hotkey('alt', 'f4')
    time.sleep(2)


def word_operation(duration):
    """Word文档操作"""
    if os.path.exists(docx_path):
        os.remove(docx_path)
    Document().save(docx_path)
    time.sleep(1)
    os.startfile(docx_path)
    time.sleep(3)

    toggle_fullscreen()
    time.sleep(2)
    set_english_keyboard()
    start_time = time.time()

    while time.time() - start_time < duration:
        pyautogui.write("To be, or not to be - that is the question.", interval=0.1)

    pyautogui.hotkey('ctrl', 's')
    time.sleep(1)
    pyautogui.hotkey('alt', 'f4')
    time.sleep(2)
    processes = ['WINWORD.EXE', 'wps.exe']
    for p in processes:
        subprocess.run(f'taskkill /F /IM {p}', shell=True)


def create_and_open_excel(duration):
    # 固定参数
    ROWS = 1000
    COLS = 5
    current_col = 6  # 起始列为F列（第六列）
    START_ROW = 4

    if os.path.exists(xlsx_path):
        os.remove(xlsx_path)

    # 创建文件并填充数据
    wb = openpyxl.Workbook()
    ws = wb.active

    # 填充5列，每列1-1000
    for col in range(1, COLS + 1):
        for row in range(1, ROWS + 1):
            ws.cell(row=row, column=col).value = row

    wb.save(xlsx_path)

    # 打开Excel
    os.startfile(xlsx_path)
    time.sleep(5)
    toggle_fullscreen()
    time.sleep(2)
    set_english_keyboard()
    # 初始化焦点
    pyautogui.click(100, 200)
    pyautogui.press('esc')
    set_english_keyboard()
    start_time = time.time()
    timeout = duration  # 转换为秒
    iteration = 0

    while time.time() - start_time < timeout:
        # 主计算循环
        try:
            # 动态生成目标列字母（如F、G等）
            target_col_letter = openpyxl.utils.get_column_letter(current_col)
            # 定位到K4
            time.sleep(1)
            pyautogui.hotkey('ctrl', 'home')
            time.sleep(0.5)
            for _ in range(current_col - 1):
                pyautogui.press('right', interval=0.1)
            for _ in range(3):
                pyautogui.press('down', interval=0.1)
            time.sleep(0.2)

            # 写入五列的求和公式到当前动态列
            for col_idx in range(COLS):
                col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                formula = f"=SUM({col_letter}1:{col_letter}{ROWS})"
                pyautogui.write(formula, interval=0.1)
                pyautogui.press('enter')
                if col_idx < COLS - 1:
                    pyautogui.press('down')
                    time.sleep(0.2)

            # 移动到下一列（如F→G）
            current_col += 1
            iteration += 1

        except Exception as e:
            print(f"出现错误: {str(e)}")
            break

    # 关闭Excel
    pyautogui.hotkey('ctrl', 's')
    time.sleep(1)
    pyautogui.hotkey('alt', 'f4')
    time.sleep(2)
    processes = ['EXCEL.EXE', 'wps.exe', 'et.exe']
    for p in processes:
        subprocess.run(f'taskkill /F /IM {p}', shell=True)


def ppt_operation(duration):
    """PPT演示操作"""
    os.startfile(ppt_path)
    time.sleep(5)
    toggle_fullscreen()
    time.sleep(2)
    pyautogui.hotkey('f5')  # 开始播放
    time.sleep(duration)
    time.sleep(2)
    pyautogui.hotkey('alt', 'f4')  # 结束播放
    time.sleep(2)
    pyautogui.hotkey('alt', 'f4')
    time.sleep(2)
    processes = ['POWERPNT.EXE', 'wpp.exe', 'wps.exe']
    for p in processes:
        subprocess.run(f'taskkill /F /IM {p}', shell=True)


def web_test(duration):
    service = Service(executable_path=edge_driver)
    options = webdriver.EdgeOptions()
    options.add_argument("--start-maximized")  # 强制初始最大化

    with webdriver.Edge(service=service, options=options) as driver:
        driver.get("https://www.bilibili.com")
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

        start_time = time.time()
        while (time.time() - start_time) < duration:
            # 滚动页面并等待
            if is_page_bottom(driver):
                driver.execute_script("window.scrollTo(0, 0);")
                time.sleep(1)
            else:
                driver.execute_script(f"window.scrollBy(0, {SW_SCROLL});")
            time.sleep(3)

            # 强制检查时间以防最后一次等待超时
            if (time.time() - start_time) >= duration:
                break


def video_operation(duration):
    options = webdriver.EdgeOptions()
    options.add_argument(f"--user-data-dir={user_data_dir}")
    options.add_argument("--profile-directory=Default")  # 使用默认配置文件
    options.add_argument("--start-maximized")
    options.add_argument("--autoplay-policy=no-user-gesture-required")

    with webdriver.Edge(service=Service(edge_driver), options=options) as driver:
        driver.get(video_url)  # 直接访问视频地址（已登录状态）
        wait = WebDriverWait(driver, 20)
        time.sleep(5)
        video_area = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".bpx-player-video-area")))
        # 双击全屏播放视频
        actions = ActionChains(driver)
        actions.move_to_element(video_area).double_click().perform()
        start_time = time.time()
        while (time.time() - start_time) < duration:
            time.sleep(0.1)

def format_duration(seconds):
    """秒转分钟格式化（保留两位小数）"""
    minutes = seconds / 60
    if minutes.is_integer():
        return f"{int(minutes)}min"
    return f"{minutes:.2f}min".rstrip('0').rstrip('.') # 智能去除多余零

def calculate_runtime(start_time_str, end_time_str, consumption, end_cap):
    """计算续航时间预估"""
    start_time = datetime.strptime(start_time_str, "%Y-%m-%d %H:%M:%S")
    end_time = datetime.strptime(end_time_str, "%Y-%m-%d %H:%M:%S")
    delta_seconds = (end_time - start_time).total_seconds()

    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate if rate > 0 else 0
        hours, remainder = divmod(remaining, 3600)
        minutes, seconds = divmod(remainder, 60)
        return f"{int(hours)}h{int(minutes)}m{int(seconds)}s"
    return "N/A"


def format_runtime(seconds):
    """统一格式化时间"""
    hours, remainder = divmod(seconds, 3600)
    minutes, sec = divmod(remainder, 60)
    return f"{int(hours)}h{int(minutes)}m{int(sec)}s"


def add_cycle_summary(cycle_num, first_start_row, current_end_row,use_run_time):
    """根据总起始行和当前结束行生成累计循环汇总"""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 新增绿色填充样式
    highlight_fill = PatternFill(start_color='C6EFCE',  # 浅绿色背景
                               end_color='C6EFCE',
                               fill_type='solid')

    # 获取总起始行数据（第一轮的第一条记录）
    first_row = ws[first_start_row]
    # 获取当前轮的最后一条记录
    last_row = ws[current_end_row]

    # 提取时间数据（需处理无年份的日期）
    start_time_str = f"2025-{first_row[1].value}"  # 假设测试年份为2025
    end_time_str = f"2025-{last_row[2].value}"
    start_time = datetime.strptime(start_time_str, "%Y-%m-%d %H:%M:%S")
    end_time = datetime.strptime(end_time_str, "%Y-%m-%d %H:%M:%S")

    # 计算时间差
    delta_seconds = (end_time - start_time).total_seconds()

    # 提取电量数据
    start_cap = first_row[5].value
    end_cap = last_row[6].value
    design_cap = first_row[7].value
    consumption = start_cap - end_cap

    # 计算续航时间（逻辑与原函数一致）
    runtime, total_runtime = "N/A", "N/A"
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate
        runtime = format_runtime(remaining)
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        total_runtime = format_runtime(total_runtime_seconds)

    # 构建汇总行数据
    summary_data = [
        f"{cycle_num}轮循环（{use_run_time}）",
        first_row[1].value,
        last_row[2].value,
        first_row[3].value,
        last_row[4].value,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    # 写入并保存
    ws.append(summary_data)
    # 应用居中样式
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = highlight_fill  # 添加背景色
    wb.save(excel_file)

def is_admin():
    try:
        return ctypes.windll.shell32.IsUserAnAdmin()
    except:
        return False

def main():
    set_english_keyboard()
    create_excel_with_header()
    # 初始化后台
    backend_preparation(5)

    first_cycle_start_row = None  # 记录总起始行

    for i in range(test_times):
        try:
            # 获取当前行号作为循环起点
            wb = openpyxl.load_workbook(excel_file)
            start_row = wb.active.max_row
            wb.close()
            # 如果是第一轮，初始化总起始行
            if first_cycle_start_row is None:
                first_cycle_start_row = start_row + 1  # 标题行后首条数据行
            time_label = format_duration(test_use_times)
            time_label_cpu_z = format_duration(cpu_z_times)
            time_label_cycle = format_duration(test_use_times*6+cpu_z_times)
            # 执行各测试步骤
            test_steps = [
                (f"网易云、QQ挂后台（{time_label}）", partial(idle_operation, duration=test_use_times)),
                (f"Word文档编辑（{time_label}）", partial(word_operation, duration=test_use_times)),
                (f"Excel表格计算（{time_label}）", partial(create_and_open_excel, duration=test_use_times)),
                (f"PPT演示播放（{time_label}）", partial(ppt_operation, duration=test_use_times)),
                (f"B站首页刷新（{time_label}）", partial(web_test, duration=test_use_times)),
                (f"B站视频播放（{time_label}）", partial(video_operation, duration=test_use_times)),
                (f"CPU-Z运行（{time_label_cpu_z}）", partial(cpu_z, duration=cpu_z_times))
            ]
            # 调整循环调用方式
            for step_name, func in test_steps:
                record_operation_step(step_name, func)

            # 获取当前循环后的结束行
            wb = openpyxl.load_workbook(excel_file)
            current_end_row = wb.active.max_row
            wb.close()

            # 添加累计循环汇总
            add_cycle_summary(
                cycle_num=i + 1,
                first_start_row=first_cycle_start_row,
                current_end_row=current_end_row,
                use_run_time=time_label_cycle
            )

        except Exception as e:
            print(f"测试中断，发生错误: {str(e)}")


if __name__ == '__main__':
    if not is_admin():
        # 重新以管理员权限运行脚本
        ctypes.windll.shell32.ShellExecuteW(None, "runas", sys.executable, " ".join(sys.argv), None, 1)
        sys.exit()
    if DEBUG_TYPE == "True":
        print(f"接下来进入环境准备，稍后会打开B站指定网页视频，请做以下操作:\n1、登录账号\n2、将视频的分辨率修改为1080P\n3、将视频的进度条移动至刚开始\n做完以上操作后可关闭网页，并在配置文件config.ini中将debug项改为False，然后重新运行程序即可", flush=True)
        time.sleep(10)
        options = webdriver.EdgeOptions()
        options.add_argument(f"--user-data-dir={user_data_dir}")
        options.add_argument("--profile-directory=Default")  # 使用默认配置文件
        options.add_argument("--start-maximized")
        options.add_argument("--autoplay-policy=no-user-gesture-required")
        with webdriver.Edge(service=Service(edge_driver), options=options) as driver:
            driver.get(video_url)  # 直接访问视频地址（已登录状态）
            time.sleep(10000)
    main()