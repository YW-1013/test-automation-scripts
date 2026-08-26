import wmi
import subprocess
import os
import time
import psutil
import openpyxl
import threading
import win32api
import win32con
import sys
import configparser
import re
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from openpyxl.styles import Alignment
from selenium.webdriver.support.ui import WebDriverWait
import ctypes

# 配置参数
current_working_dir = os.path.dirname(os.path.realpath(sys.argv[0]))
config = configparser.ConfigParser()
config.read(os.path.join(current_working_dir, 'config.ini'), encoding='utf-8')

# 路径配置
paths = config['Paths']
web_path = paths.get('web_path')
edge_driver = os.path.join(current_working_dir, 'edgedriver_win64', 'msedgedriver.exe')

# 参数配置
settings = config['Settings']
test_use_times = settings.getint('test_use_times')
test_times = settings.getint('test_times')

# 全局变量
excel_file = None
web_list = [url.strip() for url in web_path.split("|") if url.strip()]
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
SW_SCROLL = 0x100
stop_event = threading.Event()
driver_list = []
driver_list_lock = threading.Lock()


def is_page_bottom(driver):
    """判断是否滚动到页面底部"""
    try:
        scroll_height = driver.execute_script("return document.body.scrollHeight")
        scroll_top = driver.execute_script("return document.documentElement.scrollTop || window.pageYOffset")
        window_height = driver.execute_script("return window.innerHeight")
        return (scroll_top + window_height) >= scroll_height
    except:
        return False


def create_excel_with_header():
    global excel_file
    now = datetime.now()
    file_name = f"win_result_{now.strftime('%Y%m%d%H%M%S')}.xlsx"
    excel_file = os.path.join(current_working_dir, file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["操作步骤", "起始时间", "结束时间", "起始电量（%）", "结束电量（%）",
               "起始电量（mWh）", "结束电量（mWh）", "最大容量（mWh）", "消耗电量（mWh）",
               "剩余电量续航时间预估", "预估总续航时间"]
    ws.append(headers)

    column_widths = {
        "A": 30, "B": 19, "C": 19, "D": 14, "E": 14,
        "F": 15, "G": 15, "H": 15, "I": 15, "J": 21, "K": 15
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = CENTER_ALIGNMENT

    wb.save(excel_file)
    return excel_file


def write_to_excel(data):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append(data)
    for cell in ws[ws.max_row]:
        cell.alignment = CENTER_ALIGNMENT
    wb.save(excel_file)


def get_battery_capacity_info():
    try:
        c = wmi.WMI(namespace="root\\wmi")
        battery_status = c.BatteryStatus()[0]
        report_path = os.path.join(current_working_dir, "battery-report.html")
        subprocess.run(f'powercfg /batteryreport /output "{report_path}"',
                       shell=True, check=True, creationflags=subprocess.CREATE_NO_WINDOW)
        with open(report_path, 'r', encoding='utf-8') as f:
            content = f.read()
            design_pattern = r'<span class="label">DESIGN CAPACITY</span></td>\s*<td>\s*([\d,]+)\s*mWh'
            full_charge_pattern = r'<span class="label">FULL CHARGE CAPACITY</span></td>\s*<td>\s*([\d,]+)\s*mWh'
            design_match = re.search(design_pattern, content)
            full_match = re.search(full_charge_pattern, content)
            full_charge = int(full_match.group(1).replace(',', '')) if full_match else 0
        return battery_status.RemainingCapacity, full_charge
    except Exception as e:
        print(f"获取电池信息错误: {str(e)}")
        return 0, 0


def get_battery_info():
    battery = psutil.sensors_battery()
    return battery.percent if battery else 0

def get_real_screen_size():
    """获取物理像素分辨率，兼容高DPI屏幕"""
    user32 = ctypes.windll.user32
    user32.SetProcessDPIAware()
    width = user32.GetSystemMetrics(0)
    height = user32.GetSystemMetrics(1)
    return width, height


def calculate_window_position(index, total):
    screen_width, screen_height = get_real_screen_size()  # 用物理分辨率
    col = 2 if total > 2 else 1
    row = (total + 1) // 2
    width = screen_width // col
    height = screen_height // row
    x = (index % col) * width + 5
    y = (index // col) * height + 5
    return x, y, width - 10, height - 10


def browser_worker(url, index, total, driver_ready_list):
    user_data_dir = os.path.join(
        os.environ['LOCALAPPDATA'],
        'Microsoft', 'Edge', f'UserData_index{index}'
    )
    global driver_list
    while not stop_event.is_set():
        # 清理旧用户数据
        try:
            os.system(f'rmdir /s /q "{user_data_dir}"')
        except:
            pass
        time.sleep(2)
        driver = None
        try:
            options = webdriver.EdgeOptions()
            options.add_argument(f"--user-data-dir={user_data_dir}")
            options.add_argument("--disable-infobars")
            options.add_argument("--force-device-scale-factor=1")
            options.add_argument("--disable-extensions")
            options.add_argument("--disable-gpu")
            x, y, width, height = calculate_window_position(index, total)
            options.add_argument(f"--window-position={x},{y}")
            options.add_argument(f"--window-size={width},{height}")

            driver = webdriver.Edge(
                service=Service(edge_driver, service_args=['--startup-timeout=30']),
                options=options
            )
            driver.set_window_size(width, height)
            driver.set_window_position(x, y)
            driver.set_page_load_timeout(30)
            driver.get(url)
            WebDriverWait(driver, 30).until(
                lambda d: d.execute_script('return document.readyState') == 'complete'
            )
            with driver_list_lock:
                driver_list[index] = driver
                driver_ready_list[index] = True
                print(f"浏览器 {index} 初始化成功，剩余待启动：{len(web_list) - sum(driver_ready_list)}")
            # 主循环：只要stop_event没设且窗口活着就sleep
            while not stop_event.is_set():
                try:
                    # 试探driver是否活着
                    driver.execute_script("return 1;")
                except Exception:
                    print(f"浏览器{index}已被关闭或崩溃，尝试重启")
                    break  # 跳出while，重新初始化
                time.sleep(1)
        except Exception as e:
            print(f"浏览器{index}初始化失败: {str(e)}，3秒后重试")
            time.sleep(3)
        finally:
            if driver:
                try:
                    driver.execute_script("window.localStorage.clear();")
                    driver.execute_script("window.sessionStorage.clear();")
                except:
                    pass
                try:
                    driver.quit()
                except:
                    pass
            try:
                os.system(f'rmdir /s /q "{user_data_dir}"')
            except:
                pass
            with driver_list_lock:
                driver_list[index] = None
                driver_ready_list[index] = False

def scroll_window(driver, duration):
    """对driver窗口滚动duration秒"""
    start = time.time()
    while time.time() - start < duration:
        try:
            if is_page_bottom(driver):
                driver.execute_script("window.scrollTo(0, 0);")
            else:
                driver.execute_script(f"window.scrollBy(0, {SW_SCROLL});")
            time.sleep(1.5)
        except Exception as e:
            print(f"滚动操作异常: {str(e)}")
            break


def format_runtime(seconds):
    hours, remainder = divmod(seconds, 3600)
    minutes, sec = divmod(remainder, 60)
    return f"{int(hours)}h{int(minutes)}m{int(sec)}s"


def main():
    create_excel_with_header()
    global driver_list
    driver_list = [None] * len(web_list)
    threads = []
    stop_event.clear()
    driver_ready_list = [False] * len(web_list)
    # 启动所有浏览器线程
    for i, url in enumerate(web_list):
        t = threading.Thread(
            target=browser_worker,
            args=(url, i, len(web_list), driver_ready_list),
            daemon=True
        )
        t.start()
        threads.append(t)

    # 等待浏览器初始化
    print("\n=== 浏览器初始化开始 ===")
    start_time = time.time()
    while True:
        with driver_list_lock:
            current_count = sum(driver_ready_list)
        if current_count == len(web_list):
            print("\n所有浏览器初始化成功！")
            break
        if time.time() - start_time > 120:
            print(f"\n错误：仅成功启动 {current_count}/{len(web_list)} 个浏览器")
            stop_event.set()
            break
        print(f"\r已启动 {current_count}/{len(web_list)} 等待中...", end="")
        time.sleep(2)

    if stop_event.is_set():
        return

    # 执行测试轮次
    for cycle in range(test_times):
        try:
            current_cycle = cycle + 1
            step_name = f"多浏览器滚动测试第{current_cycle}轮（{test_use_times}s/每窗口）"

            round_start_time = datetime.now()
            start_percent = get_battery_info()
            start_cap, design_cap = get_battery_capacity_info()

            # 依次轮流滚动每个driver
            for idx, driver in enumerate(driver_list):
                if driver is None:
                    print(f"警告：第{idx + 1}个窗口未活跃，跳过滚动")
                    continue
                print(f"第{current_cycle}轮-第{idx+1}个窗口正在进行滚动浏览")
                # 可选：激活窗口
                try:
                    driver.switch_to.window(driver.current_window_handle)
                except Exception:
                    pass
                scroll_window(driver, test_use_times)

            round_end_time = datetime.now()
            end_percent = get_battery_info()
            end_cap, _ = get_battery_capacity_info()
            delta_seconds = (round_end_time - round_start_time).total_seconds()
            consumption = start_cap - end_cap
            runtime, total_runtime = "N/A", "N/A"

            if delta_seconds > 0 and consumption > 0:
                rate = consumption / delta_seconds
                remaining = end_cap / rate
                runtime = format_runtime(remaining)
                total_runtime_seconds = (design_cap / consumption) * delta_seconds
                total_runtime = format_runtime(total_runtime_seconds)
            data = [
                step_name,
                round_start_time.strftime("%m-%d %H:%M:%S"),
                round_end_time.strftime("%m-%d %H:%M:%S"),
                start_percent,
                end_percent,
                start_cap,
                end_cap,
                design_cap,
                consumption,
                runtime,
                total_runtime
            ]
            write_to_excel(data)

        except Exception as e:
            print(f"第{current_cycle}轮测试异常: {str(e)}")

    # 测试结束清理
    stop_event.set()
    for t in threads:
        t.join(10)
    os.system("taskkill /f /im msedge.exe /t")
    os.system("taskkill /f /im msedgewebview2.exe /t")


if __name__ == '__main__':
    main()