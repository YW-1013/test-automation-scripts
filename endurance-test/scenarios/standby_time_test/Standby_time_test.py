"""
一、多场景负载模拟
1、网易云音乐后台播放、飞书后台登录----ok
2、wps三件套docx打字、xlsx数字求和、pptx幻灯片放映  3*10分钟----ok
3、打开京东、淘宝、新浪、网易、搜狐、it之家、百度进行搜索并浏览 10分钟----ok
4、开启弹幕，观看B站 4K视频 10分钟----ok
5、飞书会议开会 10分钟----ok
6、飞书群聊天 10分钟----ok
二、记录笔记本性能数据
1、CPU占用
2、GPU占用
3、内存占用
4、电池电量
5、功耗
6、温度
7、CPU频率
"""
import ctypes
import subprocess
from docx import Document
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
import psutil
import GPUtil
import wmi
import configparser
import sys
import openpyxl
import os
import win32api
import win32process
import pyautogui
import time
import ctypes
import win32gui
import win32com.client

# 窗口状态常量
SW_SHOWMAXIMIZED = 3
SW_SHOWNORMAL = 1
SW_SHOWMINIMIZED = 2

# 定义需要的常量和函数
SW_FORCEMINIMIZE = 11
user32 = ctypes.windll.user32
# 定义键码
VK_LWIN = 0x5B  # 左 Windows 键
KEYEVENTF_KEYUP = 0x0002

current_working_dir = os.path.dirname(os.path.realpath(sys.argv[0])) # 当前工作目录
config_path = os.path.join(current_working_dir, 'config.ini')
config = configparser.ConfigParser()
config.read(config_path)
paths = config['Paths']
urls = config['Urls']
settings = config['Settings']
wangyiyun_path = paths.get('wangyiyun_path')
feishu_path = paths.get('feishu_path')
docx_path = paths.get('docx_path')
xlsx_path = paths.get('xlsx_path')
ppt_path = paths.get('ppt_path')
bilibili_url = urls.get('bilibili_url')
chrome_driver = paths.get('chrome_driver')
chrome_user_data = paths.get('chrome_user_data')

test_use_times = settings.getint('test_use_times')
row_number = settings.getint('row_number')
column_number = settings.getint('column_number')
excel_sum = settings.getint('excel_sum')
text_copy_times = settings.getint('text_copy_times')
log_path = os.path.join(current_working_dir, 'system_status_log.txt')
image_dir = os.path.join(current_working_dir, 'images')

#模拟按下win键
def press_windows_key():
    # 使用 ctypes 调用 keybd_event 函数模拟按键
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, 0, 0)  # 按下 Windows 键
    time.sleep(0.1)  # 保持按下状态
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)  # 释放 Windows 键
    time.sleep(1)
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, 0, 0)  # 按下 Windows 键
    time.sleep(0.1)  # 保持按下状态
    ctypes.windll.user32.keybd_event(VK_LWIN, 0, KEYEVENTF_KEYUP, 0)  # 释放 Windows 键

#获取当前输入法
def get_current_input_language():
    hwnd = win32gui.GetForegroundWindow()
    thread_id, _ = win32process.GetWindowThreadProcessId(hwnd)
    layout_id = win32api.GetKeyboardLayout(thread_id)

    # 提取前16位语言标识符（LANGID）
    language_code = layout_id & 0xFFFF

    # 检查语言代码的前几位，英语通常是 0x09，中文是 0x04（繁体、简体）
    if (language_code & 0x0FF) == 0x09:  # 检查是否为英语
        print("English")
        return "English"
    elif (language_code & 0x0FF) == 0x04:  # 检查是否为中文
        print("Chinese")
        return "Chinese"
    else:
        print("Unknown")
        return "Unknown"



#切换当前输入法
def ensure_english_input():
    # 通常使用 Alt + Shift 或 Windows + 空格 来切换输入法
    # 这里的例子是通过 Alt + Shift 切换输入法，具体方法可以根据你的环境调整。
    pyautogui.hotkey('alt', 'shift')
    time.sleep(2)


def is_maximized():
    hwnd = win32gui.GetForegroundWindow()
    if hwnd:
        # 定义 WINDOWPLACEMENT 结构体
        class WINDOWPLACEMENT(ctypes.Structure):
            _fields_ = [("length", ctypes.c_uint),
                        ("flags", ctypes.c_uint),
                        ("showCmd", ctypes.c_uint),
                        ("ptMinPosition", ctypes.wintypes.POINT),
                        ("ptMaxPosition", ctypes.wintypes.POINT),
                        ("rcNormalPosition", ctypes.wintypes.RECT)]

        # 创建一个 WINDOWPLACEMENT 对象并初始化其长度
        placement = WINDOWPLACEMENT()
        placement.length = ctypes.sizeof(WINDOWPLACEMENT)

        # 获取窗口显示状态
        ctypes.windll.user32.GetWindowPlacement(hwnd, ctypes.byref(placement))

        # 检查 showCmd 是否为最大化状态
        if placement.showCmd == SW_SHOWMAXIMIZED:
            return True
    return False

#全屏操作
def toggle_fullscreen():
    if not is_maximized():
        pyautogui.hotkey('win', 'up')  # 只有在当前不处于全屏时，才尝试进行全屏操作

# 最小化所有窗口
def minimize_all_windows():
    # 发送 Win + D 快捷键两次，可以最小化然后还原所有窗口
    user32.keybd_event(0x5B, 0, 0, 0)  # Press Left Windows
    user32.keybd_event(0x44, 0, 0, 0)  # Press D
    user32.keybd_event(0x44, 0, 2, 0)  # Release D
    user32.keybd_event(0x5B, 0, 2, 0)  # Release Left Windows

#启动网易云音乐播放音乐和飞书挂在后台----主运行
def Backend_application_test():

    # 启动网易云音乐应用
    subprocess.Popen(wangyiyun_path)

    #等待网易云完全启动
    time.sleep(10)
    toggle_fullscreen()
    time.sleep(5)

    # 模拟按下F5键,
    pyautogui.press('f5')
    time.sleep(5)

    #打开飞书，后台挂着
    subprocess.Popen(feishu_path)

    #等待飞书完全打开
    time.sleep(5)
    toggle_fullscreen()
    time.sleep(5)

    #最小化网易云音乐和飞书
    minimize_all_windows()
    time.sleep(10)

def repeat_text():
    text = "bairiyishanjin ,huangheruhailiu .yuqiongqianlimu ,gengshangyicenglou ."
    # 将给定的文本复制指定次数
    repeated_text = text * text_copy_times
    return repeated_text

def create_and_open_docx():
    # 检查文件是否存在，如果存在则删除文件
    if os.path.exists(docx_path):
        os.remove(docx_path)
        print(f"已删除存在的文件: {docx_path}")

    # 创建新文档
    document = Document()

    # 保存文档
    document.save(docx_path)
    print(f"新文档已保存到: {docx_path}")

    # 使用WPS Office或Microsoft Word打开文档
    try:
        if os.path.exists(docx_path):
            # 使用默认关联的应用程序打开文档
            subprocess.run(['start', docx_path], shell=True)
            print(f"文档已使用默认关联程序打开: {docx_path}")
        else:
            print(f"文件路径不存在: {docx_path}")
    except Exception as e:
        print(f"打开文档时出错: {e}")

#打开docx文档自动输入文字----主运行
def automate_typing():
    create_and_open_docx()
    # 等待几秒钟以确保文档已经打开
    time.sleep(5)
    toggle_fullscreen()
    time.sleep(5)
    ensure_english_input()


    # 使用 pyautogui 模拟输入
    pyautogui.typewrite(repeat_text(), interval=0.1)
    print("文本已自动输入完成")

    # 模拟保存操作
    # 根据不同的操作系统，快捷键组合可能不同，以下快捷键是Windows操作系统上常用的
    pyautogui.hotkey('ctrl', 's')
    time.sleep(1)  # 等待保存完成

    # 模拟关闭操作
    pyautogui.hotkey('alt', 'f4')
    time.sleep(5)  # 等待关闭完成

#打开xlsx表格进行自动计算----主运行
def create_and_open_excel(file_name, row_number, column_number, excel_sum):
    # Create a new Excel file with the specified name
    wb = openpyxl.Workbook()
    ws = wb.active

    # Fill 4 columns with data from 1 to 1000
    for col in range(1, row_number):
        for row in range(1, column_number+1):
            ws.cell(row=row, column=col).value = row

    wb.save(file_name)

    start_col = row_number - 1
    start_row = 4

    # Open the file using the default program
    os.startfile(file_name)
    time.sleep(5)  # Give enough time for Excel to open
    toggle_fullscreen()
    time.sleep(5)
    ensure_english_input()


    # Move to the specified start cell
    pyautogui.click(100,
                    200)  # Click somewhere to focus on Excel, you may adjust coordinates based on your screen resolution
    pyautogui.press('esc')  # Deselect any selected cell
    time.sleep(1)

    # Using the arrow keys or the Home key combo to move to the specific starting cell
    pyautogui.hotkey('ctrl', 'home')  # Move to A1
    time.sleep(1)
    for _ in range(start_col):
        pyautogui.press('right')
        time.sleep(0.1)
    for _ in range(start_row):
        pyautogui.press('down')
        time.sleep(0.1)

    for i in range(excel_sum):
        for j in range(row_number-1):
            col_letter = openpyxl.utils.get_column_letter(j + 1)
            sum_formula = f"=SUM({col_letter}1:{col_letter}{column_number})"

            # Type the sum formula and confirm it
            pyautogui.typewrite(sum_formula, interval=0.1)
            pyautogui.press('enter')
            time.sleep(1)

        # Move back to the starting row of the current iteration
        pyautogui.press('up', presses=row_number-1)
        time.sleep(0.1)
        # Move to the next column
        pyautogui.press('right')
        time.sleep(0.1)

    # Save final result and close the file
    pyautogui.hotkey('ctrl', 's')
    time.sleep(1)
    pyautogui.hotkey('alt', 'f4')
    time.sleep(5)

def bring_window_to_front(hwnd):
    shell = win32com.client.Dispatch("WScript.Shell")
    shell.SendKeys('%')  # 发送 Alt 键，用于解除当前可能的对话框阻挡
    win32gui.SetForegroundWindow(hwnd)

#打开PPT进行自动放映----主运行
def open_and_play_ppt_with_wps(ppt_path):
    if not os.path.exists(ppt_path):
        raise FileNotFoundError(f"PPT 文件路径不存在: {ppt_path}")

    # 使用默认程序打开PPT文件（假设WPS Office是默认的PPT处理程序）
    subprocess.Popen(['start', ppt_path], shell=True)

    # 等待PPT文件完全打开
    time.sleep(5)
    toggle_fullscreen()
    time.sleep(5)


    # 模拟按键 'F5' 开始幻灯片放映
    pyautogui.hotkey('f5')

    # 等待放映一段时间（假设30秒）
    time.sleep(test_use_times)

    # 关闭WPS应用（假设WPS已经被激活）
    pyautogui.hotkey('alt', 'f4')
    time.sleep(5)
    pyautogui.hotkey('alt', 'f4')
    time.sleep(5)



def browse_website(url, browse_time=10, scroll_interval=5):
    service = Service(chrome_driver)  # 请确保路径正确
    driver = webdriver.Chrome(service=service)
    driver.get(url)
    toggle_fullscreen()
    time.sleep(5)

    total_time = 0
    while total_time < browse_time:
        time.sleep(scroll_interval)
        driver.execute_script("window.scrollBy(0, window.innerHeight);")
        total_time += scroll_interval

    driver.quit()


def search_and_browse_baidu(search_term, total_browse_time, scroll_interval):
    service = Service(chrome_driver)  # 请确保路径正确
    driver = webdriver.Chrome(service=service)
    driver.get("https://www.baidu.com")
    toggle_fullscreen()
    time.sleep(5)

    # 搜索指定文字
    search_box = driver.find_element(By.NAME, 'wd')
    search_box.send_keys(search_term)
    search_box.send_keys(Keys.RETURN)

    time.sleep(5)  # 等待搜索结果加载

    # 点击第一个搜索结果
    first_result = driver.find_element(By.XPATH, '//div[@id="content_left"]//h3/a')
    first_result.click()

    # 切换到新打开的标签页
    driver.switch_to.window(driver.window_handles[-1])

    # 总浏览时间和滚动间隔时间
    total_time = 0
    while total_time < total_browse_time:
        time.sleep(scroll_interval)
        driver.execute_script("window.scrollBy(0, window.innerHeight);")
        total_time += scroll_interval

    driver.quit()

#打开网页进行自动浏览----主运行
def web_test():
    websites = [
        ("https://www.jd.com", 15, 5),
        ("https://www.tmall.com", 15, 5),
        ("https://www.sina.com.cn", 15, 5),
        ("https://www.sohu.com", 15, 5),
        ("https://www.163.com", 15, 5),
        ("https://www.ithome.com", 15, 5)
    ]

    for url, browse_time, scroll_interval in websites:
        browse_website(url, browse_time, scroll_interval)

    search_and_browse_baidu("巴黎奥运会", 100, 10)
    time.sleep(5)

def play_bilibili_video(video_url, test_use_times):  # 默认播放时间设为120秒（2分钟）
    service = Service(chrome_driver)  # 请确保路径正确

    driver = webdriver.Chrome(service=service)

    try:
        # 1. 打开B站指定网址
        driver.get(video_url)

        # 2. 将浏览器窗口最大化
        driver.maximize_window()

        wait = WebDriverWait(driver, 20)
        actions = ActionChains(driver)
        time.sleep(5)

        # 3. 等待视频区域加载完成
        video_area = wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".bpx-player-video-area")))

        # 播放1分钟10秒
        time.sleep(70)

        # 关闭登录弹窗
        login_close_button = driver.find_element(By.CSS_SELECTOR, '.bili-mini-mask .bili-mini-close-icon')
        if login_close_button:
            login_close_button.click()

        # 按空格键继续播放视频
        video_area.click()
        actions.send_keys(' ').perform()

        # 双击全屏播放视频
        actions.move_to_element(video_area).double_click().perform()

        # 播放剩余时间
        time.sleep(test_use_times - 70)

    except Exception as e:
        print(e)

    finally:
        driver.quit()
    time.sleep(5)

def open_app(app_path):
    try:
        # 启动应用程序
        process = subprocess.Popen(app_path)
        return process
    except Exception as e:
        print(f"无法启动应用程序: {e}")
        return None


def click_button(image_path, confidence=0.8):
    try:
        location = pyautogui.locateCenterOnScreen(image_path, confidence=confidence)
        if location is not None:
            pyautogui.click(location)
            time.sleep(1)  # 等待界面刷新
        else:
            print(f"未找到按钮: {image_path}")
    except Exception as e:
        print(f"点击按钮时发生错误: {e}")

#打开飞书进行会议十分钟----主运行
def feishu_metting():

    # 启动飞书应用程序
    app_process = open_app(feishu_path)

    toggle_fullscreen()
    time.sleep(5)


    if app_process:
        time.sleep(5)  # 等待应用启动

        video_meeting_button = os.path.join(image_dir, "video_meeting_button.jpg")
        start_meeting_button = os.path.join(image_dir, "start_meeting_button.jpg")
        camera_button = os.path.join(image_dir, "camera_button.jpg")
        microphone_button = os.path.join(image_dir, "microphone_button.jpg")
        start_meeting_now_button = os.path.join(image_dir, "start_meeting_now_button.jpg")
        end_meeting_button = os.path.join(image_dir, "end_meeting_button.jpg")



        # 点击视频会议按钮
        click_button(video_meeting_button)  # 替换为视频会议按钮的截图路径
        time.sleep(1)
        # 点击发起会议按钮
        click_button(start_meeting_button)  # 替换为发起会议按钮的截图路径
        time.sleep(1)
        # 点击摄像头按钮
        click_button(camera_button)  # 替换为摄像头按钮的截图路径
        time.sleep(1)
        # 点击麦克风按钮
        click_button(microphone_button)  # 替换为麦克风按钮的截图路径
        time.sleep(1)
        # 点击开始会议按钮
        click_button(start_meeting_now_button)  # 替换为开始会议按钮的截图路径
        time.sleep(1)
        # 等待十分钟
        time.sleep(test_use_times)

        # 点击结束会议按钮
        click_button(end_meeting_button)  # 替换为结束会议按钮的截图路径
        time.sleep(5)


def type_message(message):
    try:
        pyautogui.write(message, interval=0.1)
        pyautogui.press('enter')
    except Exception as e:
        print(f"输入消息时发生错误: {e}")

#打开飞书发送消息十分钟----主运行
def feishu_message():

    message_module_button = os.path.join(image_dir, "message_module.jpg")
    group_chat_button = os.path.join(image_dir, "group_chat_test.jpg")

    # 启动飞书应用程序
    app_process = open_app(feishu_path)


    if app_process:
        time.sleep(5)  # 等待应用启动
        toggle_fullscreen()
        time.sleep(5)
        # 点击消息模块按钮
        click_button(message_module_button)

        # 点击指定群聊“test”按钮
        click_button(group_chat_button)

        end_time = time.time() + test_use_times  # 10分钟后的时间戳

        while time.time() < end_time:
            # 输入文字消息
            type_message("Hello, this is an automated message. ")
            pyautogui.hotkey('enter')
            # 间隔3秒再发送下一条消息
            time.sleep(3)
        time.sleep(5)
        pyautogui.hotkey('alt', 'f4')
        time.sleep(5)

def get_cpu_usage():
    return psutil.cpu_percent(interval=1)

def get_gpu_usage():
    gpus = GPUtil.getGPUs()
    if gpus:
        return gpus[0].load * 100  # 获取第一个GPU的占用率
    return 0

def get_memory_usage():
    memory = psutil.virtual_memory()
    return memory.percent

def get_battery_info():
    battery = psutil.sensors_battery()
    if battery:
        return battery.percent
    return "No Battery Found"

def get_system_power_draw():
    try:
        # 获取整体功耗 (暂未实现，需要特定硬件支持)
        return "N/A"
    except:
        return "N/A"

def get_motherboard_temp():
    try:
        c = wmi.WMI(namespace="root/wmi")
        temperature_info = c.MSAcpi_ThermalZoneTemperature()
        for temp in temperature_info:
            return temp.CurrentTemperature / 10.0 - 273.15  # 转换为摄氏度
    except wmi.x_wmi as e:
        print(f"Error getting motherboard temperature: {e.com_error}")
        return "N/A"

def get_cpu_frequency():
    freq = psutil.cpu_freq()
    if freq:
        return freq.current
    return "N/A"

def record_system_info(message, log_file=log_path):

    with open(log_file, "a") as file:
        file.write("时间, CPU占用, CPU占用, 内存占用, 电量剩余百分比, 整体功耗, 主板温度, CPU频率\n")

        current_time = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        cpu_usage = get_cpu_usage()
        gpu_usage = get_gpu_usage()
        memory_usage = get_memory_usage()
        battery_percent = get_battery_info()
        power_draw = get_system_power_draw()
        motherboard_temp = get_motherboard_temp()
        cpu_freq = get_cpu_frequency()

        log_entry = f"{current_time}, CPU：{cpu_usage}, GPU：{gpu_usage}, 内存：{memory_usage}, 电量：{battery_percent}, 功耗：{power_draw}, 温度：{motherboard_temp}, CPU频率：{cpu_freq}\n"
        print(log_entry.strip())  # 打印到控制台（方便调试）
        file.write(f"{message}\n{log_entry}")
        file.flush()  # 确保内容写入文件

def main():
    # if get_current_input_language() == "Chinese":
    #     ensure_english_input()
    # record_system_info("打开网易云和飞书置于后台", log_file=log_path)
    # Backend_application_test()
    try:
        while True:
            # record_system_info("打开word进行编辑", log_file=log_path)
            # automate_typing()
            # record_system_info("打开excel进行计算", log_file=log_path)
            # create_and_open_excel(xlsx_path, row_number, column_number, excel_sum)
            # press_windows_key()
            # record_system_info("打开PPT进行播放", log_file=log_path)
            # open_and_play_ppt_with_wps(ppt_path)
            # record_system_info("打开网页进行浏览", log_file=log_path)
            # web_test()
            record_system_info("打开B站进行观看", log_file=log_path)
            play_bilibili_video(bilibili_url, test_use_times)
            # record_system_info("打开飞书会议", log_file=log_path)
            # feishu_metting()
            # record_system_info("打开飞书发消息", log_file=log_path)
            # feishu_message()
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == '__main__':
    main()