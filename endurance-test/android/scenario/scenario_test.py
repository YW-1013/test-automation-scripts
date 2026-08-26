import os
import subprocess
import time
from datetime import datetime
import sys
import uiautomator2 as u2
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from functools import partial
import configparser
import pyautogui
import win32gui
import ctypes

CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')


current_working_dir = os.path.dirname(os.path.realpath(sys.argv[0])) # 当前工作目录
docx_local_path = os.path.join(current_working_dir,'test.docx')
video_path = os.path.join(current_working_dir,'test.mp4')

config = configparser.ConfigParser()
config.read(os.path.join(current_working_dir, 'config.ini'), encoding='utf-8')

paths = config['Paths']
weixin_path = paths.get("weixin_path")

settings = config['Settings']
test_times = settings.getint('test_times')
word_time = settings.getint('word_time')
web_time = settings.getint('web_time')
weixin_time = settings.getint('weixin_time')
video_time = settings.getint('video_time')
jianying_time = settings.getint('jianying_time')

locations = config['Locations']
jianying_clickx = locations.getint('jianying_clickx')
jianying_clicky = locations.getint('jianying_clicky')
weixin_click1x = locations.getint('weixin_click1x')
weixin_click1y = locations.getint('weixin_click1y')
weixin_click2x = locations.getint('weixin_click2x')
weixin_click2y = locations.getint('weixin_click2y')
weixin_click3x = locations.getint('weixin_click3x')
weixin_click3y = locations.getint('weixin_click3y')
weixin_phone_clickx = locations.getint('weixin_phone_clickx')
weixin_phone_clicky = locations.getint('weixin_phone_clicky')

video_url = config['Urls'].get('video_url')
ip = config['Urls'].get('ip')

os.system(f"adb connect {ip} ")
device = u2.connect()
os.system(f"adb -s {ip} shell setprop persist.h3c.root_state 123@qwe")
os.system(f"adb -s {ip} root")

# 窗口状态常量
SW_SHOWMAXIMIZED = 3
width, height = device.window_size()
center_x, center_y = width // 2, height // 2
excel_file = None

def create_excel_with_header():
    global excel_file
    now = datetime.now()
    file_name = f"android_result_{now.strftime('%Y%m%d%H%M%S')}.xlsx"
    excel_file = os.path.join(current_working_dir, file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["操作步骤", "起始时间", "结束时间", "起始电量（%）", "结束电量（%）",
               "起始电量（mWh）", "结束电量（mWh）", "最大容量（mWh）", "消耗电量（mWh）", "剩余电量续航时间预估", "预估总续航时间"]
    ws.append(headers)

    # 列宽配置字典（单位：字符）
    column_widths = {
        "A": 30,  # 操作步骤
        "B": 15,  # 起始时间
        "C": 15,  # 结束时间
        "D": 14,  # 起始电量%
        "E": 14,  # 结束电量%
        "F": 15,  # 起始容量
        "G": 15,  # 结束容量
        "H": 15,  # 设计容量
        "I": 15,  # 消耗电量
        "J": 30,  # 剩余电量续航时间预估
        "K": 15  # 预估总续航
    }

    # 应用列宽
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置居中样式
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = CENTER_ALIGNMENT

    wb.save(excel_file)
    return excel_file


def write_to_excel(data):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append(data)

    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
    wb.save(excel_file)

def format_runtime(seconds):
    """统一格式化时间"""
    hours, remainder = divmod(seconds, 3600)
    minutes, sec = divmod(remainder, 60)
    return f"{int(hours)}h{int(minutes)}m{int(sec)}s"

def add_cycle_summary(cycle_num, first_start_row, current_end_row,use_run_time):
    """根据总起始行和当前结束行生成累计循环汇总"""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 新增绿色填充样式
    highlight_fill = PatternFill(start_color='C6EFCE',  # 浅绿色背景
                               end_color='C6EFCE',
                               fill_type='solid')

    # 获取总起始行数据（第一轮的第一条记录）
    first_row = ws[first_start_row]
    # 获取当前轮的最后一条记录
    last_row = ws[current_end_row]

    # 提取时间数据（需处理无年份的日期）
    start_time_str = f"2025-{first_row[1].value}"  # 假设测试年份为2025
    end_time_str = f"2025-{last_row[2].value}"
    start_time = datetime.strptime(start_time_str, "%Y-%m-%d %H:%M:%S")
    end_time = datetime.strptime(end_time_str, "%Y-%m-%d %H:%M:%S")

    # 计算时间差
    delta_seconds = (end_time - start_time).total_seconds()

    # 提取电量数据
    start_cap = first_row[5].value
    end_cap = last_row[6].value
    design_cap = first_row[7].value
    consumption = start_cap - end_cap

    # 计算续航时间（逻辑与原函数一致）
    runtime, total_runtime = "N/A", "N/A"
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate
        runtime = format_runtime(remaining)
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        total_runtime = format_runtime(total_runtime_seconds)

    # 构建汇总行数据
    summary_data = [
        f"{cycle_num}轮循环（{use_run_time}）",
        first_row[1].value,
        last_row[2].value,
        first_row[3].value,
        last_row[4].value,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    # 写入并保存
    ws.append(summary_data)
    # 应用居中样式
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = highlight_fill  # 添加背景色
    wb.save(excel_file)


def is_maximized():
    hwnd = win32gui.GetForegroundWindow()
    if hwnd:
        # 定义 WINDOWPLACEMENT 结构体
        class WINDOWPLACEMENT(ctypes.Structure):
            _fields_ = [("length", ctypes.c_uint),
                        ("flags", ctypes.c_uint),
                        ("showCmd", ctypes.c_uint),
                        ("ptMinPosition", ctypes.wintypes.POINT),
                        ("ptMaxPosition", ctypes.wintypes.POINT),
                        ("rcNormalPosition", ctypes.wintypes.RECT)]

        # 创建一个 WINDOWPLACEMENT 对象并初始化其长度
        placement = WINDOWPLACEMENT()
        placement.length = ctypes.sizeof(WINDOWPLACEMENT)

        # 获取窗口显示状态
        ctypes.windll.user32.GetWindowPlacement(hwnd, ctypes.byref(placement))

        # 检查 showCmd 是否为最大化状态
        if placement.showCmd == SW_SHOWMAXIMIZED:
            return True
    return False


#全屏操作
def toggle_fullscreen():
    if not is_maximized():
        pyautogui.hotkey('win', 'up')  # 只有在当前不处于全屏时，才尝试进行全屏操作

def send_keyevent(key_code):
    """发送ADB键盘事件"""
    subprocess.run(f'adb shell input keyevent {key_code}', shell=True)

def is_app_installed(package_name):
    """检查某个包是否已经安装"""
    try:
        result = subprocess.run(['adb', 'shell', 'pm', 'list', 'packages', package_name],
                                capture_output=True, text=True, check=True)
        return package_name in result.stdout
    except subprocess.CalledProcessError:
        return False

def check_environment():
    """检查设备上是否安装了微信、剪映和WPS"""
    apps = {
        '微信': 'com.tencent.mm',
        '剪映': 'com.lemon.lv',
        'WPS': 'cn.wps.moffice_eng'
    }

    all_installed = True
    for app_name, package_name in apps.items():
        if not is_app_installed(package_name):
            all_installed = False

    if not all_installed:
        print("环境未配置好")
    else:
        print("所有应用均已安装")

def open_and_type_text(duration):
    device.app_stop_all()
    print("执行文档操作")
    time.sleep(3)
    try:
        subprocess.run(
            ["adb", "shell", "rm", "-f", "/sdcard/AAA/test.docx"],  # -f 参数强制删除不报错
            capture_output=True,
            text=True
        )
        subprocess.run(["adb", "shell", "mkdir", "/sdcard/AAA"], check=True)
    except Exception:
        pass
    subprocess.run(["adb", "push", docx_local_path, "/sdcard/AAA"], capture_output=True, text=True)
    """通过文件管理器打开文档"""
    # 启动文件管理器（不同设备可能需要调整包名）
    device.app_start("com.huawei.filemanager")
    time.sleep(2)

    # 点击显示内部存储（根据实际UI调整）
    if device(text="内部存储").exists(timeout=5):
        device(text="内部存储").click()
    elif device(description="显示根目录").exists(timeout=5):
        device(description="显示根目录").click()

    # 导航到AAA目录
    if device(text="AAA").exists(timeout=10):
        device(text="AAA").click()
    elif device(textContains="AAA").exists(timeout=10):
        device(textContains="AAA").click()
    else:
        raise Exception("找不到AAA目录")

    # 点击目标文件
    filename = os.path.basename("/sdcard/AAA/test.docx")
    if device(text=filename).exists(timeout=10):
        device(text=filename).click()
    else:
        raise Exception(f"找不到文件 {filename}")

    if device(text="WPS Office").exists(timeout=8):
        device(text="WPS Office").click()
        # 处理可能出现的"仅此一次"按钮
        if device(text="仅此一次").exists(timeout=5):
            device(text="仅此一次").click()
        time.sleep(3)

    # 等待文档加载
    time.sleep(5)
    start_time = time.time()
    while time.time() - start_time < duration:
        # 点击屏幕坐标 (x, y)（需提前获取文本框位置）
        # device.click(center_x, center_y)

        # 或通过ADB输入
        # device.shell("am broadcast -a ADB_INPUT_TEXT --es msg 'To be, or not to be - that is the question'")
        device.shell(f"To be, or not to be - that is the question")
        time.sleep(0.5)

    # 定位输入区域
    # input_field = device(className="android.widget.EditText")
    # if input_field.wait(exists=True, timeout=20):
    #     input_field.click()
    #
    #     start_time = time.time()
    #     while time.time() - start_time < duration:
    #         input_field.set_text("To be, or not to be - that is the question.")
    #         time.sleep(0.5)
    # else:
    #     raise Exception("未找到输入区域，请检查文档是否成功打开")
    for _ in range(3):
        device.press("back")
        time.sleep(1)
    device.app_stop("com.huawei.hsl")
    device.app_stop("com.huawei.filemanager")


def auto_bilibili_browse(duration):
    print("执行B站首页刷新")
    try:
        # 获取屏幕尺寸
        w, h = device.window_size()
        # 1. 启动浏览器
        os.system(f"adb -s {ip} shell am start -a android.intent.action.VIEW -d https://www.bilibili.com")
        time.sleep(5)  # 等待页面加载
        # 3. 自动刷新循环
        start_time = datetime.now()
        while (datetime.now() - start_time).total_seconds() < duration:
            # 模拟向下滑动（屏幕Y轴80%位置滑动到20%位置）
            device.swipe(w * 0.5, h * 0.8, w * 0.5, h * 0.2, 0.5)
            time.sleep(3)  # 刷新间隔

        # 4. 关闭页面流程
        for _ in range(3):
            device.press("back")
            time.sleep(1)
        device.app_stop("com.huawei.browser")
    except Exception as e:
        print(f"执行异常: {str(e)}")

def weixin_chat(duration):
    # 静默模式参数配置
    silent_args = {
        'stdout': subprocess.DEVNULL,  # 禁止标准输出
        'stderr': subprocess.DEVNULL,  # 禁止错误输出
        'shell': True
    }
    subprocess.Popen(weixin_path, **silent_args)  # 打开微信
    time.sleep(5)
    toggle_fullscreen()  # 微信全屏
    time.sleep(5)
    pyautogui.click(x=weixin_click1x, y=weixin_click1y)#点击指定人员头像
    time.sleep(5)
    pyautogui.click(x=weixin_click2x, y=weixin_click2y)#点击视频聊天
    time.sleep(5)
    device.click(weixin_phone_clickx,weixin_phone_clicky)#手机上点击接电话
    time.sleep(duration)
    pyautogui.click(x=weixin_click3x, y=weixin_click3y)#点击结束聊天
    time.sleep(2)

def video_operation(duration):
    device.app_stop_all()
    time.sleep(3)
    try:
        subprocess.run(
            ["adb", "shell", "rm", "-f", "/sdcard/AAA/test.mp4"],  # -f 参数强制删除不报错
            capture_output=True,
            text=True
        )
        subprocess.run(["adb", "shell", "mkdir", "/sdcard/AAA"], check=True)
    except Exception:
        pass
    subprocess.run(["adb", "push", video_path, "/sdcard/AAA"], capture_output=True, text=True)
    device.app_start("com.huawei.filemanager")
    time.sleep(2)

    # 点击显示内部存储（根据实际UI调整）
    if device(text="内部存储").exists(timeout=5):
        device(text="内部存储").click()
    elif device(description="显示根目录").exists(timeout=5):
        device(description="显示根目录").click()

    # 导航到AAA目录
    if device(text="AAA").exists(timeout=10):
        device(text="AAA").click()
    elif device(textContains="AAA").exists(timeout=10):
        device(textContains="AAA").click()
    else:
        raise Exception("找不到AAA目录")
    device(text="test.mp4").click(timeout=2)
    try:
        device(text="音视频播放器").click()
    except:
        pass
    while device(text = "循环播放").exists() is False:
        device.click(center_x,center_y)
        device(text="单个播放").click()
        time.sleep(3)
    time.sleep(duration)



def jianying_operation(duration):
    device.app_stop_all()
    time.sleep(3)
    os.system(f"adb -s {ip} shell am start com.lemon.lv/com.vega.main.MainActivity")
    time.sleep(10)
    device.click(jianying_clickx,jianying_clicky)
    time.sleep(10)
    device(text = "导出").click(timeout=5)
    time.sleep(duration)


def get_battery_info():
    """通过adb获取安卓设备电池信息并计算容量"""
    result = subprocess.run(['adb', 'shell', 'cat /sys/class/power_supply/BAT0/uevent'],
                            capture_output=True, text=True)
    battery_data = {}
    for line in result.stdout.split('\n'):
        if '=' in line:
            key, value = line.split('=', 1)
            battery_data[key] = value.strip()

    # 单位换算关键参数

    voltage_design = int(battery_data.get('POWER_SUPPLY_VOLTAGE_MIN_DESIGN', 0))  # 设计电压（微伏）
    # voltage_now = int(battery_data.get('POWER_SUPPLY_VOLTAGE_NOW', 0))  # 当前电压（微伏）
    charge_full = int(battery_data.get('POWER_SUPPLY_CHARGE_FULL', 0))  # 满电量（微安时）
    charge_now = int(battery_data.get('POWER_SUPPLY_CHARGE_NOW', 0))  # 当前电量（微安时）
    capacity = int(battery_data.get('POWER_SUPPLY_CAPACITY', 0))  # 电量百分比

    # 转换为mWh单位（与Windows单位统一）
    max_capacity = (charge_full * voltage_design) / 1e9  # (μAh * μV) / 1e9 = mWh
    current_capacity = (charge_now * voltage_design) / 1e9

    return {
        'percent': capacity,
        'current_mwh': current_capacity,
        'max_mwh': max_capacity,
        'design_mwh': (int(battery_data.get('POWER_SUPPLY_CHARGE_FULL_DESIGN', 0)) * voltage_design) / 1e9
    }


def record_operation_step(step_name, operation_func, duration=None):
    """记录操作步骤的核心函数（安卓版）"""
    # 记录开始数据
    start_time = datetime.now()
    start_info = get_battery_info()
    start_percent = start_info['percent']
    start_cap = start_info['current_mwh']
    design_cap = start_info['design_mwh']

    # 执行实际操作
    if duration:
        operation_func(duration)
    else:
        operation_func()

    # 记录结束数据
    end_time = datetime.now()
    end_info = get_battery_info()
    end_percent = end_info['percent']
    end_cap = end_info['current_mwh']
    max_cap = end_info['max_mwh']

    # 计算指标
    delta_seconds = (end_time - start_time).total_seconds()
    consumption = start_cap - end_cap

    # 续航时间计算
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate if rate > 0 else 0
        hours, remainder = divmod(remaining, 3600)
        minutes, seconds = divmod(remainder, 60)
        runtime = f"{int(hours)}h{int(minutes)}m{int(seconds)}s"

        # 总续航时间（基于设计容量）
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        t_hours, t_remainder = divmod(total_runtime_seconds, 3600)
        t_minutes, t_seconds = divmod(t_remainder, 60)
        total_runtime = f"{int(t_hours)}h{int(t_minutes)}m{int(t_seconds)}s"
    else:
        runtime = "N/A"
        total_runtime = "N/A"

    # 写入Excel（保持与Windows相同的数据结构）
    data = [
        step_name,
        start_time.strftime("%m-%d %H:%M:%S"),
        end_time.strftime("%m-%d %H:%M:%S"),
        start_percent,
        end_percent,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    write_to_excel(data)

def format_duration(seconds):
    """秒转分钟格式化（保留两位小数）"""
    minutes = seconds / 60
    if minutes.is_integer():
        return f"{int(minutes)}min"
    return f"{minutes:.2f}min".rstrip('0').rstrip('.') + "min"  # 智能去除多余零

def idle_operation(duration):
    print(f"等待{duration}S")
    """静止等待操作"""
    time.sleep(duration)


def main():
    create_excel_with_header()
    #判断应用是否存在，不存在则提示用户安装
    check_environment()

    first_cycle_start_row = None  # 记录总起始行


    for i in range(test_times):
        try:
            # 获取当前行号作为循环起点
            wb = openpyxl.load_workbook(excel_file)
            start_row = wb.active.max_row
            wb.close()
            # 如果是第一轮，初始化总起始行
            if first_cycle_start_row is None:
                first_cycle_start_row = start_row + 1  # 标题行后首条数据行
            time_label_word = format_duration(word_time)
            time_label_web = format_duration(web_time)
            time_label_weixin = format_duration(weixin_time)
            time_label_video = format_duration(video_time)
            time_label_jianying = format_duration(jianying_time)

            time_label_cycle = format_duration(word_time+web_time+weixin_time+video_time+jianying_time)

            # 执行各测试步骤
            test_steps = [
                # (f"word打字（{time_label_word}）", partial(open_and_type_text, duration=word_time)),
                # (f"网页浏览（{time_label_web}）", partial(auto_bilibili_browse, duration=web_time)),
                # (f"微信视频（{time_label_weixin}）", partial(weixin_chat, duration=weixin_time)),
                (f"4K视频播放（{time_label_video}）", partial(video_operation, duration=video_time)),
                (f"剪映输出（{time_label_jianying}）", partial(jianying_operation, duration=jianying_time))
            ]
            # 调整循环调用方式
            for step_name, func in test_steps:
                record_operation_step(step_name, func)
            # 获取当前循环后的结束行
            wb = openpyxl.load_workbook(excel_file)
            current_end_row = wb.active.max_row
            wb.close()
            # 添加累计循环汇总
            add_cycle_summary(
                cycle_num=i + 1,
                first_start_row=first_cycle_start_row,
                current_end_row=current_end_row,
                use_run_time=time_label_cycle
            )
        except Exception as e:
            print(f"测试中断，发生错误: {str(e)}")

if __name__ == '__main__':
    main()