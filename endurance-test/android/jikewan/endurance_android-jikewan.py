import os
import subprocess
import time
from datetime import datetime
import sys
import uiautomator2 as u2
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from functools import partial
import configparser

CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')


current_working_dir = os.path.dirname(os.path.realpath(sys.argv[0])) # 当前工作目录

config = configparser.ConfigParser()
config.read(os.path.join(current_working_dir, 'config.ini'), encoding='utf-8')
settings = config['Settings']
test_times = settings.getint('test_times')
video_time = settings.getint('video_time') * 60
game_time = settings.getint('game_time') * 60
web_time = settings.getint('web_time') * 60
game_name = settings.get('game_name')
video_url = config['Urls'].get('video_url')
web_url = config['Urls'].get('web_url')
ip = config['Urls'].get('ip')

os.system(f"adb connect {ip} ")
device = u2.connect(ip)
os.system(f"adb -s {ip} shell setprop persist.h3c.root_state 123@qwe")
os.system(f"adb -s {ip} root")

width, height = device.window_size()
center_x, center_y = width // 2, height // 2
excel_file = None

def create_excel_with_header():
    global excel_file
    now = datetime.now()
    file_name = f"android_result_{now.strftime('%Y%m%d%H%M%S')}.xlsx"
    excel_file = os.path.join(current_working_dir, file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["操作步骤", "起始时间", "结束时间", "起始电量（%）", "结束电量（%）",
               "起始电量（mWh）", "结束电量（mWh）", "最大容量（mWh）", "消耗电量（mWh）", "剩余电量续航时间预估", "预估总续航时间"]
    ws.append(headers)

    # 列宽配置字典（单位：字符）
    column_widths = {
        "A": 30,  # 操作步骤
        "B": 15,  # 起始时间
        "C": 15,  # 结束时间
        "D": 14,  # 起始电量%
        "E": 14,  # 结束电量%
        "F": 15,  # 起始容量
        "G": 15,  # 结束容量
        "H": 15,  # 设计容量
        "I": 15,  # 消耗电量
        "J": 15,  # 可续航时间
        "K": 15  # 预估总续航
    }

    # 应用列宽
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置居中样式
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = CENTER_ALIGNMENT

    wb.save(excel_file)
    return excel_file


def write_to_excel(data):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append(data)

    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
    wb.save(excel_file)

def format_runtime(seconds):
    """统一格式化时间"""
    hours, remainder = divmod(seconds, 3600)
    minutes, sec = divmod(remainder, 60)
    return f"{int(hours)}h{int(minutes)}m{int(sec)}s"

def add_cycle_summary(cycle_num, first_start_row, current_end_row,use_run_time):
    """根据总起始行和当前结束行生成累计循环汇总"""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 新增绿色填充样式
    highlight_fill = PatternFill(start_color='C6EFCE',  # 浅绿色背景
                               end_color='C6EFCE',
                               fill_type='solid')

    # 获取总起始行数据（第一轮的第一条记录）
    first_row = ws[first_start_row]
    # 获取当前轮的最后一条记录
    last_row = ws[current_end_row]

    # 提取时间数据（需处理无年份的日期）
    start_time_str = f"2025-{first_row[1].value}"  # 假设测试年份为2025
    end_time_str = f"2025-{last_row[2].value}"
    start_time = datetime.strptime(start_time_str, "%Y-%m-%d %H:%M:%S")
    end_time = datetime.strptime(end_time_str, "%Y-%m-%d %H:%M:%S")

    # 计算时间差
    delta_seconds = (end_time - start_time).total_seconds()

    # 提取电量数据
    start_cap = first_row[5].value
    end_cap = last_row[6].value
    design_cap = first_row[7].value
    consumption = start_cap - end_cap

    # 计算续航时间（逻辑与原函数一致）
    runtime, total_runtime = "N/A", "N/A"
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate
        runtime = format_runtime(remaining)
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        total_runtime = format_runtime(total_runtime_seconds)

    # 构建汇总行数据
    summary_data = [
        f"{cycle_num}轮循环（{use_run_time/60}min）",
        first_row[1].value,
        last_row[2].value,
        first_row[3].value,
        last_row[4].value,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    # 写入并保存
    ws.append(summary_data)
    # 应用居中样式
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = highlight_fill  # 添加背景色
    wb.save(excel_file)


def auto_bilibili_video(duration):
    print("执行B站视频播放")
    device.app_stop_all()
    time.sleep(3)
    try:
        # 1. 启动浏览器
        os.system(f"adb -s {ip} shell am start -a android.intent.action.VIEW -d {video_url}")
        device(text="确定").click(timeout=5)
        time.sleep(duration)
        # 4. 关闭页面流程
        for _ in range(3):
            device.press("back")
            time.sleep(1)
        device.app_stop("mark.via")
    except Exception as e:
        print(f"执行异常: {str(e)}")

def auto_game(duration):
    if game_name == "崩铁":
        print("执行崩铁")
        device.app_stop_all()
        time.sleep(3)
        try:
            # 1. 启动浏览器
            os.system(f"adb -s {ip} shell am start com.miHoYo.hkrpg/com.mihoyo.combosdk.ComboSDKActivity")
            time.sleep(60)
            os.system(f'adb -s {ip} shell input tap {center_x} {center_y}')
            time.sleep(duration)
            # 4. 关闭页面流程
            for _ in range(3):
                device.press("back")
                if device(text="确认").exists():
                    device(text="确认").click()
                time.sleep(1)
            device.app_stop("com.miHoYo.hkrpg")
        except Exception as e:
            print(f"执行异常: {str(e)}")
    else:
        print("执行原神")
        os.system("adb shell input keyevent 3")
        time.sleep(3)
        try:
            # 1. 启动浏览器
            os.system(f"adb -s {ip} shell am start com.miHoYo.Yuanshen/com.miHoYo.GetMobileInfo.MainActivity")
            time.sleep(180)
            os.system(f'adb -s {ip} shell input tap {center_x} {center_y}')
            time.sleep(duration)
            # 4. 关闭页面流程
            for _ in range(3):
                device.press("back")
                if device(text="确认").exists():
                    device(text="确认").click()
                time.sleep(1)
            device.app_stop("com.miHoYo.Yuanshen")
        except Exception as e:
            print(f"执行异常: {str(e)}")


def run_it_home(duration):
    print("执行it之家首页浏览")
    device.app_stop_all()
    time.sleep(3)
    try:
        # 1. 启动浏览器
        w, h = device.window_size()
        os.system(f"adb -s {ip} shell am start -a android.intent.action.VIEW -d {web_url}")
        time.sleep(5)  # 等待页面加载
        start_time = datetime.now()
        while (datetime.now() - start_time).total_seconds() < duration:
            # 模拟向下滑动（屏幕Y轴80%位置滑动到20%位置）
            device.swipe(w * 0.5, h * 0.8, w * 0.5, h * 0.2, 0.5)
            time.sleep(3)  # 刷新间隔
        # 4. 关闭页面流程
        for _ in range(3):
            device.press("back")
            time.sleep(1)
        device.app_stop("mark.via")
    except Exception as e:
        print(f"执行异常: {str(e)}")

def get_battery_level():
    # 执行ADB命令获取电池信息
    result = subprocess.run(['adb', '-s', f'{ip}', 'shell', 'dumpsys', 'battery'], capture_output=True, text=True)
    output = result.stdout

    # 解析电量百分比
    level = None
    for line in output.splitlines():
        if 'level' in line.lower():
            level = int(line.split(':')[1].strip())
            break
    return level

def get_battery_info():
    """通过adb获取安卓设备电池信息并计算容量"""
    result = subprocess.run(['adb', '-s', f'{ip}', 'shell', 'cat /sys/class/power_supply/BAT0/uevent'],
                            capture_output=True, text=True)
    battery_data = {}
    for line in result.stdout.split('\n'):
        if '=' in line:
            key, value = line.split('=', 1)
            battery_data[key] = value.strip()

    # 单位换算关键参数

    voltage_design = int(battery_data.get('POWER_SUPPLY_VOLTAGE_MIN_DESIGN', 0))  # 设计电压（微伏）
    # voltage_now = int(battery_data.get('POWER_SUPPLY_VOLTAGE_NOW', 0))  # 当前电压（微伏）
    charge_full = int(battery_data.get('POWER_SUPPLY_CHARGE_FULL', 0))  # 满电量（微安时）
    charge_now = int(battery_data.get('POWER_SUPPLY_CHARGE_NOW', 0))  # 当前电量（微安时）
    capacity = int(get_battery_level())  # 电量百分比

    # 转换为mWh单位（与Windows单位统一）
    max_capacity = (charge_full * voltage_design) / 1e9  # (μAh * μV) / 1e9 = mWh
    current_capacity = (charge_now * voltage_design) / 1e9

    return {
        'percent': capacity,
        'current_mwh': current_capacity,
        'max_mwh': max_capacity,
        'design_mwh': (int(battery_data.get('POWER_SUPPLY_CHARGE_FULL_DESIGN', 0)) * voltage_design) / 1e9
    }


def record_operation_step(step_name, operation_func, duration=None):
    """记录操作步骤的核心函数（安卓版）"""
    # 记录开始数据
    start_time = datetime.now()
    start_info = get_battery_info()
    start_percent = start_info['percent']
    start_cap = start_info['current_mwh']
    design_cap = start_info['design_mwh']

    # 执行实际操作
    if duration:
        operation_func(duration)
    else:
        operation_func()

    # 记录结束数据
    end_time = datetime.now()
    end_info = get_battery_info()
    end_percent = end_info['percent']
    end_cap = end_info['current_mwh']
    max_cap = end_info['max_mwh']

    # 计算指标
    delta_seconds = (end_time - start_time).total_seconds()
    consumption = start_cap - end_cap

    # 续航时间计算
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate if rate > 0 else 0
        hours, remainder = divmod(remaining, 3600)
        minutes, seconds = divmod(remainder, 60)
        runtime = f"{int(hours)}h{int(minutes)}m{int(seconds)}s"

        # 总续航时间（基于设计容量）
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        t_hours, t_remainder = divmod(total_runtime_seconds, 3600)
        t_minutes, t_seconds = divmod(t_remainder, 60)
        total_runtime = f"{int(t_hours)}h{int(t_minutes)}m{int(t_seconds)}s"
    else:
        runtime = "N/A"
        total_runtime = "N/A"

    # 写入Excel（保持与Windows相同的数据结构）
    data = [
        step_name,
        start_time.strftime("%m-%d %H:%M:%S"),
        end_time.strftime("%m-%d %H:%M:%S"),
        start_percent,
        end_percent,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    write_to_excel(data)

def format_duration(seconds):
    """秒转分钟格式化（保留两位小数）"""
    minutes = seconds / 60
    if minutes.is_integer():
        return f"{int(minutes)}min"
    return f"{minutes:.2f}min".rstrip('0').rstrip('.') + "min"  # 智能去除多余零

def idle_operation(duration):
    print(f"等待{duration}S")
    """静止等待操作"""
    time.sleep(duration)


def main():
    create_excel_with_header()

    first_cycle_start_row = None  # 记录总起始行


    for i in range(test_times):
        try:
            # 获取当前行号作为循环起点
            wb = openpyxl.load_workbook(excel_file)
            start_row = wb.active.max_row
            wb.close()
            # 如果是第一轮，初始化总起始行
            if first_cycle_start_row is None:
                first_cycle_start_row = start_row + 1  # 标题行后首条数据行
            # 执行各测试步骤
            test_steps = [
                (f"B站视频播放（30min）", partial(auto_bilibili_video, duration=video_time)),
                (f"游戏运行（20min）", partial(auto_game, duration=game_time)),
                (f"IT之家浏览（10min）", partial(run_it_home, duration=web_time)),
            ]
            # 调整循环调用方式
            for step_name, func in test_steps:
                record_operation_step(step_name, func)
            # 获取当前循环后的结束行
            wb = openpyxl.load_workbook(excel_file)
            current_end_row = wb.active.max_row
            wb.close()
            # 添加累计循环汇总
            add_cycle_summary(
                cycle_num=i + 1,
                first_start_row=first_cycle_start_row,
                current_end_row=current_end_row,
                use_run_time=video_time + game_time + web_time
            )
        except Exception as e:
            print(f"测试中断，发生错误: {str(e)}")

if __name__ == '__main__':
    main()