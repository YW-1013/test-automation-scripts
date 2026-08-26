import os
import subprocess
import time
from datetime import datetime
import sys
import uiautomator2 as u2
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from functools import partial
import configparser
import gc


CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')


current_working_dir = os.path.dirname(os.path.realpath(sys.argv[0])) # 当前工作目录
docx_local_path = os.path.join(current_working_dir,'test.docx')
excel_local_path = os.path.join(current_working_dir,'test.xlsx')
ppt_local_path = os.path.join(current_working_dir,'test.pptx')

config = configparser.ConfigParser()
config.read(os.path.join(current_working_dir, 'config.ini'), encoding='utf-8')
settings = config['Settings']
test_times = settings.getint('test_times')
test_use_times = settings.getint('test_use_times')
video_url = config['Urls'].get('video_url')
ip = config['Urls'].get('ip')
excel_file = None

def cleanup_device(device):
    """统一清理设备资源"""
    try:
        device.app_stop_all()  # 停止所有应用
        device.service("uiautomator").stop()  # 停止uiautomator服务
    except:
        pass
    time.sleep(1)


def create_excel_with_header():
    global excel_file
    now = datetime.now()
    file_name = f"android_result_{now.strftime('%Y%m%d%H%M%S')}.xlsx"
    excel_file = os.path.join(current_working_dir, file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["操作步骤", "起始时间", "结束时间", "起始电量（%）", "结束电量（%）",
               "起始电量（mWh）", "结束电量（mWh）", "最大容量（mWh）", "消耗电量（mWh）", "剩余电量续航时间预估", "预估总续航时间"]
    ws.append(headers)

    # 列宽配置字典（单位：字符）
    column_widths = {
        "A": 30,  # 操作步骤
        "B": 15,  # 起始时间
        "C": 15,  # 结束时间
        "D": 14,  # 起始电量%
        "E": 14,  # 结束电量%
        "F": 15,  # 起始容量
        "G": 15,  # 结束容量
        "H": 15,  # 设计容量
        "I": 15,  # 消耗电量
        "J": 15,  # 可续航时间
        "K": 15  # 预估总续航
    }

    # 应用列宽
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置居中样式
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = CENTER_ALIGNMENT

    wb.save(excel_file)
    return excel_file


def write_to_excel(data):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append(data)
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
    wb.save(excel_file)  # 保存后自动关闭

def format_runtime(seconds):
    """统一格式化时间"""
    hours, remainder = divmod(seconds, 3600)
    minutes, sec = divmod(remainder, 60)
    return f"{int(hours)}h{int(minutes)}m{int(sec)}s"

def add_cycle_summary(cycle_num, first_start_row, current_end_row,use_run_time):
    """根据总起始行和当前结束行生成累计循环汇总"""
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 新增绿色填充样式
    highlight_fill = PatternFill(start_color='C6EFCE',  # 浅绿色背景
                               end_color='C6EFCE',
                               fill_type='solid')

    # 获取总起始行数据（第一轮的第一条记录）
    first_row = ws[first_start_row]
    # 获取当前轮的最后一条记录
    last_row = ws[current_end_row]

    # 提取时间数据（需处理无年份的日期）
    start_time_str = f"2025-{first_row[1].value}"  # 假设测试年份为2025
    end_time_str = f"2025-{last_row[2].value}"
    start_time = datetime.strptime(start_time_str, "%Y-%m-%d %H:%M:%S")
    end_time = datetime.strptime(end_time_str, "%Y-%m-%d %H:%M:%S")

    # 计算时间差
    delta_seconds = (end_time - start_time).total_seconds()

    # 提取电量数据
    start_cap = first_row[5].value
    end_cap = last_row[6].value
    design_cap = first_row[7].value
    consumption = start_cap - end_cap

    # 计算续航时间（逻辑与原函数一致）
    runtime, total_runtime = "N/A", "N/A"
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate
        runtime = format_runtime(remaining)
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        total_runtime = format_runtime(total_runtime_seconds)

    # 构建汇总行数据
    summary_data = [
        f"{cycle_num}轮循环（{use_run_time}）",
        first_row[1].value,
        last_row[2].value,
        first_row[3].value,
        last_row[4].value,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    # 写入并保存
    ws.append(summary_data)
    # 应用居中样式
    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
        cell.fill = highlight_fill  # 添加背景色
    wb.save(excel_file)



def send_keyevent(key_code):
    """发送ADB键盘事件"""
    subprocess.run(f'adb shell input keyevent {key_code}', shell=True)

def is_app_installed(package_name):
    """检查某个包是否已经安装"""
    try:
        result = subprocess.run(['adb', 'shell', 'pm', 'list', 'packages', package_name],
                                capture_output=True, text=True, check=True)
        return package_name in result.stdout
    except subprocess.CalledProcessError:
        return False

def check_environment():
    """检查设备上是否安装了QQ、网易云音乐和WPS"""
    apps = {
        'QQ': 'com.tencent.mobileqq',
        '网易云音乐': 'com.netease.cloudmusic',
        'WPS': 'cn.wps.moffice_eng',
        '哔哩哔哩': 'tv.danmaku.bili'
    }

    all_installed = True
    for app_name, package_name in apps.items():
        if not is_app_installed(package_name):
            all_installed = False

    if not all_installed:
        print("环境未配置好")
    else:
        print("所有应用均已安装")

def execute_adb(cmd):
    """执行adb命令并确保进程退出"""
    proc = subprocess.Popen(cmd, shell=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    stdout, stderr = proc.communicate()
    proc.terminate()
    return stdout.decode().strip()


def open_music_qq(device):
    # device.app_stop_all()
    execute_adb("adb shell input keyevent 3")
    time.sleep(2)
    print("打开网易云和QQ置于后台")
    """打开网易云音乐并点击播放按钮"""
    device.app_start("com.netease.cloudmusic")  # 启动网易云音乐应用
    time.sleep(10)  # 等待应用加载

    # 假设播放按钮的资源 ID 为 'play_button'，需要根据实际 UI 结构调整
    try:
        play_button = device(resourceId="com.netease.cloudmusic:id/minPlayBtn")
        if play_button.exists(timeout=30):
            play_button.click()
        else:
            print("未找到播放按钮，确认是否在正确的页面")
    except Exception as e:
        print(f"播放按钮点击失败: {str(e)}")
    device.app_start("com.tencent.mobileqq")  # 启动QQ应用
    time.sleep(5)  # 等待应用加载

def open_and_type_text(duration):
    print("执行文档操作")
    device.app_stop("com.h3c.filemanager")
    time.sleep(1)
    execute_adb("adb shell input keyevent 3")
    time.sleep(3)
    # 检查文件是否存在
    file_exists_cmd = f"adb -s {ip} shell ls /sdcard/AAA/test.docx"
    result = execute_adb(file_exists_cmd)
    if result != 0:
        try:
            execute_adb(f"adb -s {ip} shell mkdir /sdcard/AAA")
            time.sleep(3)
            execute_adb(f"adb -s {ip} shell rm -f /sdcard/AAA/test.docx")

        except Exception:
            pass
        time.sleep(3)
        execute_adb(f"adb -s {ip} push {docx_local_path} /sdcard/AAA")
        time.sleep(3)
    """通过文件管理器打开文档"""
    # 启动文件管理器（不同设备可能需要调整包名）
    device.app_start("com.h3c.filemanager")
    time.sleep(2)

    # 导航到AAA目录
    if device(text="AAA").exists(timeout=10):
        device(text="AAA").click()
    elif device(textContains="AAA").exists(timeout=10):
        device(textContains="AAA").click()
    else:
        print("找不到AAA目录")

    # 点击目标文件
    filename = os.path.basename("/sdcard/AAA/test.docx")
    if device(text=filename).exists(timeout=10):
        device(text=filename).click()
    else:
        raise Exception(f"找不到文件 {filename}")

    if device(text="WPS Office").exists(timeout=8):
        device(text="WPS Office").click()
        # 处理可能出现的"仅此一次"按钮
        if device(text="仅此一次").exists(timeout=5):
            device(text="仅此一次").click()
        time.sleep(3)

    # 等待文档加载
    time.sleep(5)

    start_time = time.time()
    while time.time() - start_time < duration:
        execute_adb("adb shell input text 'To be, or not to be - that is the question.'")
        time.sleep(0.5)
    for _ in range(5):
        device.press("back")
        time.sleep(1)
        if device(text="不保存").exists():
            device(text="不保存").click()
            device.app_stop("com.h3c.filemanager")
            break
    device.app_stop("com.h3c.filemanager")

def open_and_type_excel(duration):
    print("执行表格操作")
    # device.app_stop_all()
    execute_adb("adb shell input keyevent 3")
    ROWS = 1000
    COLS = 5
    current_col = 6  # 起始列为F列（第六列）
    START_ROW = 4
    time.sleep(3)
    file_exists_cmd = f"adb -s {ip} shell ls /sdcard/AAA/test.docx"
    result = execute_adb(file_exists_cmd)
    if result != 0:
        try:
            execute_adb(f"adb -s {ip} shell mkdir /sdcard/AAA")
            execute_adb(f"adb -s {ip} shell rm -f /sdcard/AAA/test.xlsx")
        except Exception:
            pass
        time.sleep(3)
        execute_adb(f"adb -s {ip} push {excel_local_path} /sdcard/AAA")
        time.sleep(5)
    """通过文件管理器打开文档"""
    # 启动文件管理器（不同设备可能需要调整包名）
    device.app_start("com.h3c.filemanager")
    time.sleep(2)

    # 导航到AAA目录
    if device(text="AAA").exists(timeout=10):
        device(text="AAA").click()
    elif device(textContains="AAA").exists(timeout=10):
        device(textContains="AAA").click()
    else:
        raise Exception("找不到AAA目录")

    # 点击目标文件
    filename = os.path.basename("/sdcard/AAA/test.xlsx")
    if device(text=filename).exists(timeout=10):
        device(text=filename).click()
    else:
        raise Exception(f"找不到文件 {filename}")

    if device(text="WPS Office").exists(timeout=8):
        device(text="WPS Office").click()
        # 处理可能出现的"仅此一次"按钮
        if device(text="仅此一次").exists(timeout=5):
            device(text="仅此一次").click()
        time.sleep(3)

    # 等待文档加载
    time.sleep(5)


    subprocess.run(f'adb shell input tap {center_x} {center_y}')
    time.sleep(1)
    device.press('enter')
    time.sleep(1)
    # 主操作循环
    start_time = time.time()
    while time.time() - start_time < duration:
        try:
            for col in range(COLS):
                editor = device(resourceId="cn.wps.moffice_eng:id/et_formula_editor")
                if not editor.exists:
                    editor = device(className="android.widget.EditText", focused=True)
                if editor.exists(timeout=3):
                    editor.set_text("=SUM(A1:A1000)")  # 直接输入未转义内容
                    time.sleep(1)
                    device.press('enter')
                    # time.sleep(1)
                    # device.press('enter')
                else:
                    print("无法定位输入框")
                    # subprocess.run(f'adb shell input tap {center_x} {center_y}')
                    time.sleep(1)
                    device.press('enter')
                    time.sleep(1)

                if col < COLS - 1:
                    time.sleep(1)
                    subprocess.run('adb shell input keyevent 19')  # UP
                    subprocess.run('adb shell input keyevent 22')  # RIGHT
                    time.sleep(0.2)

                # 纵向移动并回到起始列
            time.sleep(1)
            subprocess.run('adb shell input keyevent 20')  # DOWN
            for _ in range(COLS - 1):
                subprocess.run('adb shell input keyevent 21')  # LEFT
                time.sleep(0.2)
        except Exception as e:
            print(f"Error: {str(e)}")
            break

    for _ in range(5):
        device.press("back")
        time.sleep(1)
        if device(text="不保存").exists():
            device(text="不保存").click()
            device.app_stop("com.h3c.filemanager")
            break
    device.app_stop("com.h3c.filemanager")

def play_ppt_with_shortcuts(duration):
    print("执行PPT自动放映")
    device.app_stop("com.h3c.filemanager")
    execute_adb("adb shell input keyevent 3")
    time.sleep(3)
    file_exists_cmd = f"adb -s {ip} shell ls /sdcard/AAA/test.docx"
    result = execute_adb(file_exists_cmd)
    if result != 0:
        try:
            execute_adb(f"adb -s {ip} shell mkdir /sdcard/AAA")
            execute_adb(f"adb -s {ip} shell rm -f /sdcard/AAA/test.pptx")
        except Exception:
            pass
        time.sleep(3)
        execute_adb(f"adb -s {ip} push {ppt_local_path} /sdcard/AAA")
        time.sleep(3)
    """通过文件管理器打开文档"""
    # 启动文件管理器（不同设备可能需要调整包名）
    device.app_start("com.h3c.filemanager")
    time.sleep(2)

    # 导航到AAA目录
    if device(text="AAA").exists(timeout=10):
        device(text="AAA").click()
    elif device(textContains="AAA").exists(timeout=10):
        device(textContains="AAA").click()
    else:
        raise Exception("找不到AAA目录")

    # 点击目标文件
    filename = os.path.basename("/sdcard/AAA/test.pptx")
    if device(text=filename).exists(timeout=10):
        device(text=filename).click()
    else:
        raise Exception(f"找不到文件 {filename}")
    if device(text="WPS Office").exists(timeout=8):
        device(text="WPS Office").click()
        # 处理可能出现的"仅此一次"按钮
        if device(text="仅此一次").exists(timeout=5):
            device(text="仅此一次").click()
        time.sleep(3)

    # 等待WPS加载完成
    device(resourceId="cn.wps.moffice_eng:id/root_view").wait(timeout=15)
    for i in range(22):
        device.press('up')
        time.sleep(0.2)

    # 快捷键操作流程
    device(text="放映").click(timeout=5)
    device(text="自动放映").click(timeout=5)
    # 开始循环播放
    time.sleep(duration)  # 每页停留时间


    for _ in range(3):
        device.press("back")
        time.sleep(1)
        if device(text="不保存").exists():
            device(text="不保存").click()
            device.app_stop("com.h3c.filemanager")
            break
    device.app_stop("com.h3c.filemanager")

def auto_bilibili_browse(duration):
    print("执行B站首页刷新")
    # device.app_stop_all()
    execute_adb("adb shell input keyevent 3")
    time.sleep(3)
    try:
        # 获取屏幕尺寸
        w, h = device.window_size()
        # 1. 启动浏览器
        execute_adb(f"adb -s {ip} shell am start -a android.intent.action.VIEW -d https://www.bilibili.com")
        time.sleep(5)  # 等待页面加载
        # 3. 自动刷新循环
        start_time = datetime.now()
        while (datetime.now() - start_time).total_seconds() < duration:
            # 模拟向下滑动（屏幕Y轴80%位置滑动到20%位置）
            device.swipe(w * 0.5, h * 0.8, w * 0.5, h * 0.2, 0.5)
            time.sleep(3)  # 刷新间隔

        # 4. 关闭页面流程
        for _ in range(3):
            device.press("back")
            time.sleep(1)
        device.app_stop("mark.via")
    except Exception as e:
        print(f"执行异常: {str(e)}")

def auto_bilibili_video(duration):
    print("执行B站视频播放")
    # device.app_stop_all()
    execute_adb("adb shell input keyevent 3")
    time.sleep(3)
    try:
        # 1. 启动浏览器
        execute_adb(f"adb -s {ip} shell am start -a android.intent.action.VIEW -d {video_url}")
        device(text="确定").click(timeout=5)
        time.sleep(duration)
        # 4. 关闭页面流程
        for _ in range(3):
            device.press("back")
            time.sleep(1)
        device.app_stop("tv.danmaku.bilibilihd")
        device.app_stop("mark.via")
    except Exception as e:
        print(f"执行异常: {str(e)}")


def get_battery_info():
    """通过adb获取安卓设备电池信息并计算容量"""
    result = subprocess.run(['adb', 'shell', 'cat /sys/class/power_supply/BAT0/uevent'],
                            capture_output=True, text=True)
    battery_data = {}
    for line in result.stdout.split('\n'):
        if '=' in line:
            key, value = line.split('=', 1)
            battery_data[key] = value.strip()

    # 单位换算关键参数

    voltage_design = int(battery_data.get('POWER_SUPPLY_VOLTAGE_MIN_DESIGN', 0))  # 设计电压（微伏）
    # voltage_now = int(battery_data.get('POWER_SUPPLY_VOLTAGE_NOW', 0))  # 当前电压（微伏）
    charge_full = int(battery_data.get('POWER_SUPPLY_ENERGY_FULL', 0))  # 满电量（微安时）
    charge_now = int(battery_data.get('POWER_SUPPLY_ENERGY_NOW', 0))  # 当前电量（微安时）
    capacity = int(battery_data.get('POWER_SUPPLY_CAPACITY', 0))  # 电量百分比

    # 转换为mWh单位（与Windows单位统一）
    max_capacity = charge_full / 1000  # (μAh * μV) / 1e9 = mWh
    current_capacity = charge_now / 1000
    return {
        'percent': capacity,
        'current_mwh': current_capacity,
        'max_mwh': max_capacity,
        'design_mwh': (int(battery_data.get('POWER_SUPPLY_ENERGY_FULL_DESIGN', 0)) * voltage_design) / 1e9
    }


def record_operation_step(step_name, operation_func, duration=None):
    """记录操作步骤的核心函数（安卓版）"""
    # 记录开始数据
    start_time = datetime.now()
    start_info = get_battery_info()
    start_percent = start_info['percent']
    start_cap = start_info['current_mwh']
    design_cap = start_info['design_mwh']

    # 执行实际操作
    if duration:
        operation_func(duration)
    else:
        operation_func()

    # 记录结束数据
    end_time = datetime.now()
    end_info = get_battery_info()
    end_percent = end_info['percent']
    end_cap = end_info['current_mwh']
    max_cap = end_info['max_mwh']

    # 计算指标
    delta_seconds = (end_time - start_time).total_seconds()
    consumption = start_cap - end_cap

    # 续航时间计算
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate if rate > 0 else 0
        hours, remainder = divmod(remaining, 3600)
        minutes, seconds = divmod(remainder, 60)
        runtime = f"{int(hours)}h{int(minutes)}m{int(seconds)}s"

        # 总续航时间（基于设计容量）
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        t_hours, t_remainder = divmod(total_runtime_seconds, 3600)
        t_minutes, t_seconds = divmod(t_remainder, 60)
        total_runtime = f"{int(t_hours)}h{int(t_minutes)}m{int(t_seconds)}s"
    else:
        runtime = "N/A"
        total_runtime = "N/A"

    # 写入Excel（保持与Windows相同的数据结构）
    data = [
        step_name,
        start_time.strftime("%m-%d %H:%M:%S"),
        end_time.strftime("%m-%d %H:%M:%S"),
        start_percent,
        end_percent,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    write_to_excel(data)

def format_duration(seconds):
    """秒转分钟格式化（保留两位小数）"""
    minutes = seconds / 60
    if minutes.is_integer():
        return f"{int(minutes)}min"
    return f"{minutes:.2f}min".rstrip('0').rstrip('.') + "min"  # 智能去除多余零

def idle_operation(duration):
    print(f"等待{duration}S")
    """静止等待操作"""
    time.sleep(duration)


def main():
    global center_x,center_y,device
    execute_adb(f"adb connect {ip} ")
    device = u2.connect(ip)
    execute_adb(f"adb -s {ip} shell setprop persist.h3c.root_state 123@qwe")
    execute_adb(f"adb -s {ip} root")

    width, height = device.window_size()
    center_x, center_y = width // 2, height // 2


    create_excel_with_header()
    #判断应用是否存在，不存在则提示用户安装
    check_environment()

    #打开qq和网易云
    open_music_qq(device)

    first_cycle_start_row = None  # 记录总起始行


    for i in range(test_times):
        # try:
        # 获取当前行号作为循环起点
        wb = openpyxl.load_workbook(excel_file)
        start_row = wb.active.max_row
        wb.close()
        # 如果是第一轮，初始化总起始行
        if first_cycle_start_row is None:
            first_cycle_start_row = start_row + 1  # 标题行后首条数据行
        time_label = format_duration(test_use_times)
        time_label_cycle = format_duration(test_use_times*6)
        # 执行各测试步骤
        test_steps = [
            (f"网易云、QQ挂后台（{time_label}）", partial(idle_operation, duration=test_use_times)),
            (f"Word文档编辑（{time_label}）", partial(open_and_type_text, duration=test_use_times)),
            (f"Excel表格计算（{time_label}）", partial(open_and_type_excel, duration=test_use_times)),
            (f"PPT演示播放（{time_label}）", partial(play_ppt_with_shortcuts, duration=test_use_times)),
            (f"B站首页刷新（{time_label}）", partial(auto_bilibili_browse, duration=test_use_times)),
            (f"B站视频播放（{time_label}）", partial(auto_bilibili_video, duration=test_use_times))
        ]
        # 调整循环调用方式
        for step_name, func in test_steps:
            record_operation_step(step_name, func)
        # 获取当前循环后的结束行
        wb = openpyxl.load_workbook(excel_file)
        current_end_row = wb.active.max_row
        wb.close()
        # 添加累计循环汇总
        add_cycle_summary(
            cycle_num=i + 1,
            first_start_row=first_cycle_start_row,
            current_end_row=current_end_row,
            use_run_time=time_label_cycle
        )
        gc.collect()
        time.sleep(2)  # 给系统释放资源的时间
        cleanup_device(device)  # 清理设备状
        execute_adb(f"adb disconnect")
        time.sleep(10)
        try:
            execute_adb(f"adb connect {ip}")
            time.sleep(10)
            device = u2.connect(ip)
            execute_adb(f"adb -s {ip} shell setprop persist.h3c.root_state 123@qwe")
            execute_adb(f"adb -s {ip} root")
        except:
            execute_adb(f"adb kill-server")
            time.sleep(10)
            execute_adb(f"adb connect {ip}")
            time.sleep(10)
            device = u2.connect(ip)
            execute_adb(f"adb -s {ip} shell setprop persist.h3c.root_state 123@qwe")
            execute_adb(f"adb -s {ip} root")

        # except Exception as e:
        #     print(f"测试中断，发生错误: {str(e)}")

if __name__ == '__main__':
    main()