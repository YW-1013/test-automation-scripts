import wmi
import subprocess
import os
import time
import psutil
import openpyxl
import sys
import configparser
from datetime import datetime
from openpyxl.styles import Alignment
import re
from openpyxl.utils import get_column_letter
import uiautomator2 as u2
from functools import partial

CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

# 配置加载
current_working_dir = os.path.dirname(os.path.realpath(sys.argv[0]))
config = configparser.ConfigParser()
config.read(os.path.join(current_working_dir, 'config.ini'), encoding='utf-8')

# 参数配置
settings = config['Settings']
time_interval = settings.getint('time_interval')
ip = config['Urls'].get('ip')
os.system(f"adb connect {ip} ")
device = u2.connect()
os.system(f"adb -s {ip} shell setprop persist.h3c.root_state 123@qwe")
os.system(f"adb -s {ip} root")

width, height = device.window_size()
center_x, center_y = width // 2, height // 2
# 全局变量
excel_file = None



def create_excel_with_header():
    global excel_file
    now = datetime.now()
    file_name = f"android_result_{now.strftime('%Y%m%d%H%M%S')}.xlsx"
    excel_file = os.path.join(current_working_dir, file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["操作步骤", "起始时间", "结束时间", "起始电量（%）", "结束电量（%）",
               "起始电量（mWh）", "结束电量（mWh）", "最大容量（mWh）", "消耗电量（mWh）", "剩余电量续航时间预估", "预估总续航时间"]
    ws.append(headers)

    # 列宽配置字典（单位：字符）
    column_widths = {
        "A": 30,  # 操作步骤
        "B": 15,  # 起始时间
        "C": 15,  # 结束时间
        "D": 14,  # 起始电量%
        "E": 14,  # 结束电量%
        "F": 15,  # 起始容量
        "G": 15,  # 结束容量
        "H": 15,  # 设计容量
        "I": 15,  # 消耗电量
        "J": 15,  # 可续航时间
        "K": 15  # 预估总续航
    }

    # 应用列宽
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置居中样式
    for row in ws.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.alignment = CENTER_ALIGNMENT

    wb.save(excel_file)
    return excel_file


def get_battery_info():
    battery = psutil.sensors_battery()
    return battery.percent if battery else 0


def get_battery_capacity_info():
    try:
        c = wmi.WMI(namespace="root\\wmi")
        battery_status = c.BatteryStatus()[0]
        script_dir = os.path.dirname(os.path.realpath(sys.argv[0]))
        report_path = os.path.join(script_dir, "battery-report.html")
        subprocess.run(f'powercfg /batteryreport /output "{report_path}"',
                       shell=True,
                       check=True,
                       creationflags=subprocess.CREATE_NO_WINDOW)
        with open(report_path, 'r', encoding='utf-8') as f:
            content = f.read()

            # 改进版正则表达式
            design_pattern = r'<span class="label">DESIGN CAPACITY</span></td>\s*<td>\s*([\d,]+)\s*mWh'
            full_charge_pattern = r'<span class="label">FULL CHARGE CAPACITY</span></td>\s*<td>\s*([\d,]+)\s*mWh'

            design_match = re.search(design_pattern, content)
            full_match = re.search(full_charge_pattern, content)

            if not design_match or not full_match:
                raise ValueError("无法在报告中找到电池容量信息")
            full_charge = int(full_match.group(1).replace(',', ''))
        return battery_status.RemainingCapacity, full_charge
    except:
        return 0, 0





def video_operation():
    device.app_stop_all()
    time.sleep(3)
    try:
        subprocess.run(
            ["adb", "shell", "rm", "-f", "/sdcard/AAA/test.mp4"],  # -f 参数强制删除不报错
            capture_output=True,
            text=True
        )
        subprocess.run(["adb", "shell", "mkdir", "/sdcard/AAA"], check=True)
    except Exception:
        pass
    device(text="AAA").click(timeout=2)
    device(text="test.wav").click(timeout=2)
    try:
        device(text="音视频播放器").click()
    except:
        pass
    while device(text = "循环播放").exists() is False:
        device.click(center_x,center_y)
        device(text="单个播放").click()
        time.sleep(3)

def idle_times():
    time.sleep(time_interval)

def get_battery_info():
    """通过adb获取安卓设备电池信息并计算容量"""
    result = subprocess.run(['adb', 'shell', 'cat /sys/class/power_supply/BAT0/uevent'],
                            capture_output=True, text=True)
    battery_data = {}
    for line in result.stdout.split('\n'):
        if '=' in line:
            key, value = line.split('=', 1)
            battery_data[key] = value.strip()

    # 单位换算关键参数

    voltage_design = int(battery_data.get('POWER_SUPPLY_VOLTAGE_MIN_DESIGN', 0))  # 设计电压（微伏）
    # voltage_now = int(battery_data.get('POWER_SUPPLY_VOLTAGE_NOW', 0))  # 当前电压（微伏）
    charge_full = int(battery_data.get('POWER_SUPPLY_CHARGE_FULL', 0))  # 满电量（微安时）
    charge_now = int(battery_data.get('POWER_SUPPLY_CHARGE_NOW', 0))  # 当前电量（微安时）
    capacity = int(battery_data.get('POWER_SUPPLY_CAPACITY', 0))  # 电量百分比

    # 转换为mWh单位（与Windows单位统一）
    max_capacity = (charge_full * voltage_design) / 1e9  # (μAh * μV) / 1e9 = mWh
    current_capacity = (charge_now * voltage_design) / 1e9

    return {
        'percent': capacity,
        'current_mwh': current_capacity,
        'max_mwh': max_capacity,
        'design_mwh': (int(battery_data.get('POWER_SUPPLY_CHARGE_FULL_DESIGN', 0)) * voltage_design) / 1e9
    }

def write_to_excel(data):
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    ws.append(data)

    last_row = ws.max_row
    for cell in ws[last_row]:
        cell.alignment = CENTER_ALIGNMENT
    wb.save(excel_file)


def record_operation_step(step_name, operation_func, duration=None):
    """记录操作步骤的核心函数（安卓版）"""
    # 记录开始数据
    start_time = datetime.now()
    start_info = get_battery_info()
    start_percent = start_info['percent']
    start_cap = start_info['current_mwh']
    design_cap = start_info['design_mwh']

    # 执行实际操作
    if duration:
        operation_func(duration)
    else:
        operation_func()

    # 记录结束数据
    end_time = datetime.now()
    end_info = get_battery_info()
    end_percent = end_info['percent']
    end_cap = end_info['current_mwh']
    max_cap = end_info['max_mwh']

    # 计算指标
    delta_seconds = (end_time - start_time).total_seconds()
    consumption = start_cap - end_cap

    # 续航时间计算
    if delta_seconds > 0 and consumption > 0:
        rate = consumption / delta_seconds
        remaining = end_cap / rate if rate > 0 else 0
        hours, remainder = divmod(remaining, 3600)
        minutes, seconds = divmod(remainder, 60)
        runtime = f"{int(hours)}h{int(minutes)}m{int(seconds)}s"

        # 总续航时间（基于设计容量）
        total_runtime_seconds = (design_cap / consumption) * delta_seconds
        t_hours, t_remainder = divmod(total_runtime_seconds, 3600)
        t_minutes, t_seconds = divmod(t_remainder, 60)
        total_runtime = f"{int(t_hours)}h{int(t_minutes)}m{int(t_seconds)}s"
    else:
        runtime = "N/A"
        total_runtime = "N/A"

    # 写入Excel（保持与Windows相同的数据结构）
    data = [
        step_name,
        start_time.strftime("%m-%d %H:%M:%S"),
        end_time.strftime("%m-%d %H:%M:%S"),
        start_percent,
        end_percent,
        start_cap,
        end_cap,
        design_cap,
        consumption,
        runtime,
        total_runtime
    ]
    write_to_excel(data)

def format_duration(seconds):
    """秒转分钟格式化（保留两位小数）"""
    minutes = seconds / 60
    if minutes.is_integer():
        return f"{int(minutes)}min"
    return f"{minutes:.2f}min".rstrip('0').rstrip('.') + "min"  # 智能去除多余零

def format_runtime(seconds):
    """统一格式化时间"""
    hours, remainder = divmod(seconds, 3600)
    minutes, sec = divmod(remainder, 60)
    return f"{int(hours)}h{int(minutes)}m{int(sec)}s"


def main():
    create_excel_with_header()


    first_cycle_start_row = None  # 记录总起始行
    times = 0
    while True:
        try:
            times += 1
            # 获取当前行号作为循环起点
            wb = openpyxl.load_workbook(excel_file)
            start_row = wb.active.max_row
            wb.close()
            # 如果是第一轮，初始化总起始行
            if first_cycle_start_row is None:
                first_cycle_start_row = start_row + 1  # 标题行后首条数据行
            time_label = format_duration(time_interval)
            video_operation()
            # 执行各测试步骤
            test_steps = [
                (f"第{times}轮测试（{time_label}）", partial(idle_times)),
            ]
            # 调整循环调用方式
            for step_name, func in test_steps:
                record_operation_step(step_name, func)
            wb.close()
        except Exception as e:
            print(f"测试中断，发生错误: {str(e)}")


if __name__ == '__main__':
    main()